# -*- coding: utf-8 -*-
"""
PDF Parser Agent (Phase 1)
v4: New Parsing Rules (Team Meeting 2026-03-19)

5 Rules agreed with team:
1. (X.Y) numbered questions -> Main CDP question row
2. Question Details table -> SKIP (not needed in Excel)
3. Number table (Number/Sub-question/Options) ->
   One Excel row per number: 번호 | Sub질문 | Options (merged, page-spanning)
4. Requested Content / Explanation of Terms -> SKIP
5. Tags table -> light gray row = 항목 (left col), white row = 내용 (right col)

Excel flat row structure per question:
  Row type "질문":   질문번호 | 질문내용 | 페이지
  Row type "번호행": 번호 | Sub질문 | Options
  Row type "태그행": Tags_항목 | Tags_내용
"""

import re
import time
import logging
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Any

HAS_PDFPLUMBER = False
try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    pass

HAS_PANDAS = False
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    pass

HAS_OPENPYXL_STYLES = False
try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL_STYLES = True
except ImportError:
    pass

from backend.core.models import (
    AgentResult, AgentStatus, ValidationResult, QuestionItem, PDFParseResult
)
from backend.core.skill_loader import load_skill
from backend.core.config import settings

try:
    from backend.agents.question_structurer import structure_questions
    HAS_STRUCTURER = True
except ImportError:
    HAS_STRUCTURER = False

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# PATTERNS
# ---------------------------------------------------------------------------
SECTION_PATTERN = re.compile(r"^\(([A-Za-z]?\d+(?:\.\d+[a-z]?)*)\)\s*(.*)")
QUESTION_ID_PATTERN = re.compile(
    r"^([A-Z]\d+\.\d+[a-z]?(?:\.\d+)?)\s+(.*)", re.MULTILINE
)
POINTS_PATTERN = re.compile(r"(\d+(?:\.\d+)?)\s*[Pp]oints?")

# Color thresholds for background classification
COLOR_DARK_GRAY = 0.80
COLOR_LIGHT_GRAY = 0.96

# Color class labels
CC_TITLE = "제목"    # dark gray header row
CC_HEADER = "헤더"   # light gray field/label row
CC_CONTENT = "내용"  # white content row


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------
def _clean_cell(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in ("nan", "none"):
        return ""
    return s


def _classify_color(r_val, g_val, b_val) -> str:
    avg = (r_val + g_val + b_val) / 3.0
    if avg <= COLOR_DARK_GRAY:
        return CC_TITLE
    elif avg <= COLOR_LIGHT_GRAY:
        return CC_HEADER
    else:
        return CC_CONTENT


def _get_page_rects(page) -> List[Dict]:
    rects = []
    raw_rects = page.rects if hasattr(page, "rects") else []
    for rect in raw_rects:
        fill = rect.get("non_stroking_color")
        if fill is None:
            continue
        if isinstance(fill, (list, tuple)):
            if len(fill) == 1:
                r_val = g_val = b_val = fill[0]
            elif len(fill) == 3:
                r_val, g_val, b_val = fill
            elif len(fill) == 4:
                c, m, y, k = fill
                r_val = (1 - c) * (1 - k)
                g_val = (1 - m) * (1 - k)
                b_val = (1 - y) * (1 - k)
            else:
                continue
        elif isinstance(fill, (int, float)):
            r_val = g_val = b_val = float(fill)
        else:
            continue

        avg_color = (r_val + g_val + b_val) / 3.0
        if avg_color < 0.1:
            continue

        rect_w = rect.get("x1", 0) - rect.get("x0", 0)
        rect_h = rect.get("bottom", 0) - rect.get("top", 0)
        if rect_w <= 3 or rect_h <= 3:
            continue

        color_class = _classify_color(r_val, g_val, b_val)
        if color_class == CC_CONTENT and r_val >= 0.99 and g_val >= 0.99 and b_val >= 0.99:
            continue

        rects.append({
            "top": rect.get("top", 0),
            "bottom": rect.get("bottom", 0),
            "x0": rect.get("x0", 0),
            "x1": rect.get("x1", 0),
            "color_class": color_class,
            "color_rgb": (round(r_val, 3), round(g_val, 3), round(b_val, 3)),
        })
    return rects


def _classify_table_by_color(
    table_data: List[List],
    table_bbox: Tuple,
    page_rects: List[Dict],
) -> List[Dict]:
    if not table_data:
        return []

    results = []
    num_rows = len(table_data)
    if num_rows == 0:
        return results

    tbl_top = table_bbox[1]
    tbl_bottom = table_bbox[3]
    tbl_height = tbl_bottom - tbl_top
    row_height = tbl_height / num_rows if num_rows > 0 else tbl_height

    for row_idx, row in enumerate(table_data):
        row_top = tbl_top + row_idx * row_height
        row_bottom = row_top + row_height
        row_mid = (row_top + row_bottom) / 2.0

        best_class = CC_CONTENT
        for rect in page_rects:
            if rect["top"] <= row_mid + 2 and rect["bottom"] >= row_mid - 2:
                if rect["x0"] < table_bbox[2] and rect["x1"] > table_bbox[0]:
                    rc = rect["color_class"]
                    if rc == CC_TITLE:
                        best_class = CC_TITLE
                        break
                    elif rc == CC_HEADER and best_class != CC_TITLE:
                        best_class = CC_HEADER

        results.append({
            "row": row,
            "color_class": best_class,
        })

    return results


# ---------------------------------------------------------------------------
# Table type detection
# ---------------------------------------------------------------------------
def _detect_table_type(classified_rows: List[Dict]) -> str:
    """
    Detect table type from first non-empty cell.
    Returns: 'question_details', 'numbered_columns', 'requested_content',
             'explanation', 'tags', or 'unknown'
    """
    first_cell = ""
    for cr in classified_rows:
        row = cr.get("row", [])
        for cell in (row or []):
            val = _clean_cell(cell)
            if val:
                first_cell = val
                break
        if first_cell:
            break

    if not first_cell:
        return "unknown"

    fl = first_cell.lower()
    if "question detail" in fl:
        return "question_details"
    if "tag" in fl:
        return "tags"
    if "requested" in fl:
        return "requested_content"
    if "explanation" in fl:
        return "explanation"
    if first_cell.strip().isdigit():
        return "numbered_columns"

    return "unknown"


def _normalize_sub_question(text: str) -> str:
    """Join multi-line sub-question text into a single line with spaces."""
    if not text:
        return text
    # Replace newlines (and surrounding whitespace) with a single space
    import re as _re
    text = _re.sub(r'\s*\n\s*', ' ', text)
    text = _re.sub(r' +', ' ', text).strip()
    return text


def _normalize_options(text: str) -> str:
    """
    Normalize Options text:
    - Bullet lines (starting with • or ●) stay on their own line.
    - Non-bullet lines are appended to the previous line with a space.
    - Continuation lines within a bullet block join the bullet with a space.
    """
    if not text:
        return text
    lines = text.split('\n')
    result = []
    for raw in lines:
        line = raw.strip()
        if not line:
            continue
        # Detect bullet by first character code point
        first_cp = ord(line[0]) if line else 0
        # 0x2022=•  0x25CF=●  0x2219=∙  0x00B7=·  0x25CB=○  0x25E6=◦
        is_bullet = first_cp in (0x2022, 0x25CF, 0x2219, 0x00B7, 0x25CB, 0x25E6)
        if is_bullet:
            result.append(line)
        else:
            if result:
                result[-1] = result[-1] + ' ' + line
            else:
                result.append(line)
    return '\n'.join(result)


# ---------------------------------------------------------------------------
# Rule 3: Extract numbered rows
# ---------------------------------------------------------------------------
def _extract_numbered_rows(classified_rows: List[Dict]) -> Tuple[List[Dict], List[Tuple]]:
    """
    Extract structured rows from a numbered columns table.
    Returns (results, col_ranges). results = list of {"번호": str, "Sub질문": str, "Options": str}

    Table structure:
    - Row with digits only (0,1,2,3...) = number header row
    - Row(s) with dark/light gray background = sub-question descriptions per column
    - Row(s) with white background = option values per column
    """
    all_rows = [
        (
            [_clean_cell(c) for c in (cr.get("row", []) or [])],
            cr.get("color_class", CC_CONTENT)
        )
        for cr in classified_rows
    ]

    if not all_rows:
        return []

    # Find the number row (row where most non-empty cells are digits)
    number_row_idx = -1
    for ri, (row_cells, cc) in enumerate(all_rows):
        non_empty = [c.strip() for c in row_cells if c.strip()]
        if not non_empty:
            continue
        digit_count = sum(1 for c in non_empty if c.isdigit())
        if digit_count >= 2 and digit_count >= len(non_empty) * 0.6:
            number_row_idx = ri
            break

    if number_row_idx < 0:
        return [], []

    num_cells = all_rows[number_row_idx][0]
    num_positions = [
        (ci, c.strip())
        for ci, c in enumerate(num_cells)
        if c.strip() and c.strip().isdigit()
    ]

    if not num_positions:
        return [], []

    col_ranges = []
    for i, (ci, num_str) in enumerate(num_positions):
        start = ci
        end = num_positions[i + 1][0] - 1 if i + 1 < len(num_positions) else len(num_cells) - 1
        col_ranges.append((num_str, start, end))

    results = []
    for num_str, start, end in col_ranges:
        # PASS 1: Find the first NON-EMPTY row for this column range.
        # This is the sub-question row (structural heuristic).
        # We skip leading empty rows because some tables have a formatting
        # spacer row between the number row and the sub-question row.
        first_data_ri = None
        for ri in range(number_row_idx + 1, len(all_rows)):
            row_cells, _ = all_rows[ri]
            for ci in range(start, min(end + 1, len(row_cells))):
                v = row_cells[ci].strip() if ci < len(row_cells) and row_cells[ci] else ""
                if v:
                    first_data_ri = ri
                    break
            if first_data_ri is not None:
                break

        sub_parts = []
        option_parts = []

        # PASS 2: Classify each row.
        for ri in range(number_row_idx + 1, len(all_rows)):
            row_cells, cc = all_rows[ri]
            range_vals = []
            for ci in range(start, min(end + 1, len(row_cells))):
                v = row_cells[ci].strip() if ci < len(row_cells) and row_cells[ci] else ""
                if v:
                    range_vals.append(v)

            if not range_vals:
                continue

            combined = " ".join(range_vals)

            if first_data_ri is not None and ri == first_data_ri:
                # First non-empty row = sub-question (always, regardless of color)
                sub_parts.append(combined)
            elif cc in (CC_TITLE, CC_HEADER):
                # Additional gray rows = sub-question continuation (rare)
                sub_parts.append(combined)
            elif (sub_parts and not option_parts
                  and len(combined) <= 8
                  and not any(c in combined for c in ("•", "●", "Select", "Please"))):
                # Short non-bullet white row right after sub-question, before any options:
                # likely a continuation of the column header split across rows (e.g. "ID")
                sub_parts.append(combined)
            else:
                # White rows = options
                option_parts.append(combined)

        results.append({
            "번호": num_str,
            "Sub질문": _normalize_sub_question(" ".join(sub_parts).strip()),
            "Options": _normalize_options("\n".join(option_parts).strip()),
        })

    return results, col_ranges


# ---------------------------------------------------------------------------
# Rule 5: Extract tags rows
# ---------------------------------------------------------------------------
def _extract_tags_rows(classified_rows: List[Dict]) -> List[Dict]:
    """
    Extract rows from a Tags table.
    Returns list of {"Tags_항목": str, "Tags_내용": str}

    Table structure:
    - Dark gray row: title row ("Tags") -> skip
    - Light gray row: 항목 (field name = left col, value = right cols)
    - White row: continuation of previous 내용
    """
    # Known CDP tag category names (fix: filter explanation text mixed into tags table)
    CDP_TAG_CATEGORIES = ["authority type", "environmental issue", "questionnaire sector"]

    results = []
    in_tags_section = False  # only process rows after "Tags" dark gray header

    for cr in classified_rows:
        cc = cr.get("color_class", CC_CONTENT)
        row = cr.get("row", [])
        cells = [_clean_cell(c) for c in (row or [])]

        if cc == CC_TITLE:
            non_empty = [c.strip() for c in cells if c.strip()]
            if non_empty and "tag" in non_empty[0].lower():
                in_tags_section = True
            continue

        if not in_tags_section:
            continue

        if cc == CC_HEADER:
            non_empty = [c.strip() for c in cells if c.strip()]
            if non_empty:
                hangmok = non_empty[0]
                hangmok_lower = hangmok.lower()
                if any(cat in hangmok_lower for cat in CDP_TAG_CATEGORIES):
                    naeong = " ".join(non_empty[1:]) if len(non_empty) > 1 else ""
                    results.append({"Tags_항목": hangmok, "Tags_내용": naeong})
        else:
            non_empty = [c.strip() for c in cells if c.strip()]
            if non_empty and results:
                additional = " ".join(non_empty)
                if results[-1]["Tags_내용"]:
                    results[-1]["Tags_내용"] += " " + additional
                else:
                    results[-1]["Tags_내용"] = additional

    return results



# ---------------------------------------------------------------------------
# Page-spanning continuation: extract options from tables without numeric row
# ---------------------------------------------------------------------------
def _extract_continuation_options(
    classified_rows: List[Dict],
    col_ranges: List[Tuple],
) -> Dict[str, str]:
    """
    Extract options from a continuation table that has no numeric header row.
    Used for page-spanning numbered tables where options overflow to the next page.

    The continuation table on the next page may:
    - Repeat sub-question headers (gray rows) as "fixed rows" -> skip these
    - Have white rows with continued option values -> extract by column position

    Args:
        classified_rows: rows from the continuation table
        col_ranges: [(num_str, start_col, end_col)] from the previous numbered table

    Returns: {num_str: options_text}
    """
    if not col_ranges:
        return {}

    options_by_num: Dict[str, str] = {}

    for cr in classified_rows:
        row_cells = [_clean_cell(c) for c in (cr.get("row", []) or [])]
        cc = cr.get("color_class", CC_CONTENT)

        # Skip gray rows (repeated sub-question header "fixed rows")
        if cc in (CC_TITLE, CC_HEADER):
            continue

        # Extract white rows by column position
        has_content = False
        for num_str, start, end in col_ranges:
            range_vals = []
            for ci in range(start, min(end + 1, len(row_cells))):
                v = row_cells[ci].strip() if ci < len(row_cells) and row_cells[ci] else ""
                if v:
                    range_vals.append(v)

            if range_vals:
                has_content = True
                new_opts = " ".join(range_vals)
                if num_str in options_by_num:
                    options_by_num[num_str] += "\n" + new_opts
                else:
                    options_by_num[num_str] = new_opts

    return options_by_num

# ---------------------------------------------------------------------------
# Build structured flat rows
# ---------------------------------------------------------------------------
def _build_structured_questions(elements: List[Dict]) -> List[Dict]:
    """
    Build flat list of rows for Excel output.

    Row types:
      {"행_유형": "질문",   "질문번호": ..., "질문내용": ..., "페이지": ...}
      {"행_유형": "번호행", "번호": ..., "Sub질문": ..., "Options": ...}
      {"행_유형": "태그행", "Tags_항목": ..., "Tags_내용": ...}

    Rules applied:
    1. (X.Y) text -> 질문 row
    2. question_details table -> SKIP
    3. numbered_columns table -> 번호행 rows (merged across pages)
    4. requested_content / explanation table -> SKIP
    5. tags table -> 태그행 rows
    """
    items = []

    for elem in elements:
        etype = elem["type"]
        page = elem["page"]
        y = elem["y_top"]

        if etype == "text":
            for line in elem["data"].split("\n"):
                line_stripped = line.strip()
                if not line_stripped:
                    continue
                m = SECTION_PATTERN.match(line_stripped)
                if m:
                    items.append((page, y, "question", {
                        "num": "(%s)" % m.group(1),
                        "text": m.group(2).strip(),
                    }))
                    y += 0.1

        elif etype == "table":
            items.append((page, y, "table", {
                "classified_rows": elem.get("classified_rows", []),
            }))

    items.sort(key=lambda x: (x[0], x[1]))

    flat_rows = []
    current_q_row = None
    pending_numbered = {}
    pending_tags = []
    last_col_ranges = []    # col ranges from last numbered table (page-spanning)

    def _flush_question():
        nonlocal current_q_row, pending_numbered, pending_tags
        if current_q_row is None:
            return
        flat_rows.append(current_q_row)

        def _num_sort_key(n):
            try:
                return int(n)
            except (ValueError, TypeError):
                return 9999

        for num_str in sorted(pending_numbered.keys(), key=_num_sort_key):
            row = pending_numbered[num_str]
            flat_rows.append({
                "행_유형": "번호행",
                "번호": num_str,
                "Sub질문": row["Sub질문"],
                "Options": row["Options"],
            })

        for tag_row in pending_tags:
            tag_row["행_유형"] = "태그행"
            flat_rows.append(tag_row)

        current_q_row = None
        pending_numbered = {}
        pending_tags = []
        last_col_ranges.clear()

    for page, y, itype, idata in items:
        if itype == "question":
            _flush_question()
            current_q_row = {
                "행_유형": "질문",
                "질문번호": idata["num"],
                "질문내용": idata["text"],
                "페이지": page,
            }
        elif itype == "table" and current_q_row is not None:
            classified_rows = idata.get("classified_rows", [])
            if not classified_rows:
                continue

            table_type = _detect_table_type(classified_rows)

            if table_type in ("question_details", "requested_content", "explanation"):
                continue

            elif table_type == "numbered_columns":
                new_rows, col_ranges = _extract_numbered_rows(classified_rows)
                if col_ranges:
                    last_col_ranges.clear()
                    last_col_ranges.extend(col_ranges)
                for nr in new_rows:
                    num = nr["번호"]
                    if num in pending_numbered:
                        if nr["Options"]:
                            if pending_numbered[num]["Options"]:
                                pending_numbered[num]["Options"] += "\n" + nr["Options"]
                            else:
                                pending_numbered[num]["Options"] = nr["Options"]
                        if not pending_numbered[num]["Sub질문"] and nr["Sub질문"]:
                            pending_numbered[num]["Sub질문"] = nr["Sub질문"]
                    else:
                        pending_numbered[num] = {
                            "Sub질문": nr["Sub질문"],
                            "Options": nr["Options"],
                        }

            elif table_type == "tags":
                new_tags = _extract_tags_rows(classified_rows)
                pending_tags.extend(new_tags)

            elif table_type == "unknown" and last_col_ranges:
                # Fix 2: page-spanning continuation table (no numeric row)
                cont_opts = _extract_continuation_options(classified_rows, last_col_ranges)
                for num_str, opts_text in cont_opts.items():
                    if num_str in pending_numbered and opts_text:
                        existing = pending_numbered[num_str]["Options"]
                        if existing:
                            pending_numbered[num_str]["Options"] = existing + "\n" + opts_text
                        else:
                            pending_numbered[num_str]["Options"] = opts_text

    _flush_question()
    return flat_rows


# ===========================================================================
# STAGE 1: Extract page elements
# ===========================================================================
def _extract_page_elements(
    pdf_path: str,
    page_start: Optional[int] = None,
    page_end: Optional[int] = None,
) -> Tuple[List[Dict], int]:
    if not HAS_PDFPLUMBER:
        raise ImportError("pdfplumber is required for PDF parsing")

    elements: List[Dict] = []
    total_pages = 0

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        start_idx = (page_start - 1) if page_start else 0
        end_idx = page_end if page_end else total_pages
        start_idx = max(0, start_idx)
        end_idx = min(total_pages, end_idx)

        for page_idx in range(start_idx, end_idx):
            page = pdf.pages[page_idx]
            page_num = page_idx + 1

            try:
                page_rects = _get_page_rects(page)

                tables = page.find_tables()
                table_bboxes = []
                for tbl in tables:
                    try:
                        bbox = tbl.bbox
                        table_bboxes.append(bbox)
                        table_data = tbl.extract()

                        classified_rows = _classify_table_by_color(
                            table_data, bbox, page_rects
                        )

                        elements.append({
                            "type": "table",
                            "page": page_num,
                            "y_top": bbox[1],
                            "data": table_data,
                            "bbox": bbox,
                            "classified_rows": classified_rows,
                        })
                    except Exception as te:
                        logger.warning("Page %d: table extraction error: %s", page_num, str(te))

                text_outside = ""
                try:
                    def not_in_table(obj, _bboxes=table_bboxes):
                        obj_y_mid = (obj.get("top", 0) + obj.get("bottom", 0)) / 2.0
                        for bbox in _bboxes:
                            if (bbox[1] - 2) <= obj_y_mid <= (bbox[3] + 2):
                                return False
                        return True

                    filtered_page = page.filter(not_in_table)
                    text_outside = filtered_page.extract_text() or ""
                except Exception:
                    text_outside = page.extract_text() or ""

                if text_outside.strip():
                    try:
                        words = filtered_page.extract_words() or []
                        q_positions = {}
                        for w in words:
                            wtext = w.get("text", "")
                            if SECTION_PATTERN.match(wtext) or (wtext.startswith("(") and wtext.endswith(")")):
                                q_positions[wtext] = w.get("top", 0)
                    except Exception:
                        q_positions = {}

                    text_lines = text_outside.split("\n")
                    current_segment = []
                    current_y = 0
                    if table_bboxes:
                        current_y = max(0, min(b[1] for b in table_bboxes) - 1)

                    for tline in text_lines:
                        tline_stripped = tline.strip()
                        if not tline_stripped:
                            continue
                        m_q = SECTION_PATTERN.match(tline_stripped)
                        if m_q:
                            if current_segment:
                                elements.append({
                                    "type": "text",
                                    "page": page_num,
                                    "y_top": current_y,
                                    "data": "\n".join(current_segment),
                                })
                                current_segment = []
                            q_key = "(%s)" % m_q.group(1)
                            current_y = q_positions.get(q_key, current_y + 0.1)
                        current_segment.append(tline_stripped)

                    if current_segment:
                        elements.append({
                            "type": "text",
                            "page": page_num,
                            "y_top": current_y,
                            "data": "\n".join(current_segment),
                        })

            except Exception as page_err:
                logger.error("Page %d processing error: %s", page_num, str(page_err))

    elements.sort(key=lambda e: (e["page"], e["y_top"]))
    return elements, total_pages


# ===========================================================================
# STAGE 2: Classify text lines
# ===========================================================================
def _classify_text_lines(text: str) -> List[Dict[str, str]]:
    results: List[Dict[str, str]] = []
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue

        m_section = SECTION_PATTERN.match(line)
        if m_section:
            results.append({
                "분류": "질문/섹션",
                "질문번호": "(%s)" % m_section.group(1),
                "내용": m_section.group(2).strip(),
            })
            continue

        m_qid = QUESTION_ID_PATTERN.match(line)
        if m_qid:
            results.append({
                "분류": "질문ID",
                "질문번호": m_qid.group(1),
                "내용": m_qid.group(2).strip(),
            })
            continue

        results.append({
            "분류": "일반",
            "질문번호": "",
            "내용": line,
        })

    return results


# ===========================================================================
# STAGE 3: Save Excel
# ===========================================================================


def _apply_structured_styles_to_sheet(writer, sheet_name: str, struct_rows: list) -> None:
    if not HAS_OPENPYXL_STYLES:
        return
    try:
        wb = writer.book
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        # Apply same header and row styling logic
        from openpyxl.styles import PatternFill, Font, Alignment
        DARK_GREEN = PatternFill("solid", fgColor="1F5C35")
        LIGHT_GREEN = PatternFill("solid", fgColor="D9EAD3")
        WHITE = PatternFill("solid", fgColor="FFFFFF")
        HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        CELL_FONT = Font(name="Calibri", size=10)
        WRAP = Alignment(wrap_text=True, vertical="top")
        # Header row
        for cell in ws[1]:
            cell.fill = DARK_GREEN
            cell.font = HDR_FONT
            cell.alignment = WRAP
        ROW_TYPE_MAP = {
            "질문": DARK_GREEN,
            "번호행": LIGHT_GREEN,
            "태그행": WHITE,
        }
        for ri, row_dict in enumerate(struct_rows, start=2):
            rt = row_dict.get("행_유형", "")
            fill = ROW_TYPE_MAP.get(rt, WHITE)
            font_color = "FFFFFF" if rt == "질문" else "000000"
            for cell in ws[ri]:
                cell.fill = fill
                cell.font = Font(name="Calibri", size=10,
                                 bold=(rt == "질문"), color=font_color)
                cell.alignment = WRAP
    except Exception as e:
        logger.warning("Style sheet error: %s", e)


def _save_faithful_excel(
    elements: List[Dict],
    output_path: str,
    source_file: str = "",
    structured_questions: list = None,
) -> Dict[str, int]:
    if not HAS_PANDAS:
        raise ImportError("pandas is required for Excel output")

    if structured_questions is None:
        flat_rows = _build_structured_questions(elements)
    else:
        flat_rows = structured_questions

    # Count tables for stats
    table_count = 0
    total_table_data_rows = 0
    for elem in elements:
        if elem["type"] == "table":
            table_count += 1
            tdata = elem.get("data") or []
            total_table_data_rows += len(tdata)

    text_row_count = sum(
        len((elem.get("data") or "").splitlines())
        for elem in elements if elem["type"] == "text"
    )

    # Build structured questions DataFrame
    sq_columns = ["\ud589_\uc720\ud615", "\uc9c8\ubb38\ubc88\ud638", "\uc9c8\ubb38\ub0b4\uc6a9",
                  "\ud398\uc774\uc9c0", "\ubc88\ud638", "Sub\uc9c8\ubb38", "Options",
                  "Tags_\ud56d\ubaa9", "Tags_\ub0b4\uc6a9"]
    struct_rows = []
    for row in flat_rows:
        struct_rows.append({
            "\ud589_\uc720\ud615": row.get("\ud589_\uc720\ud615", ""),
            "\uc9c8\ubb38\ubc88\ud638": row.get("\uc9c8\ubb38\ubc88\ud638", ""),
            "\uc9c8\ubb38\ub0b4\uc6a9": row.get("\uc9c8\ubb38\ub0b4\uc6a9", ""),
            "\ud398\uc774\uc9c0": row.get("\ud398\uc774\uc9c0", ""),
            "\ubc88\ud638": row.get("\ubc88\ud638", ""),
            "Sub\uc9c8\ubb38": row.get("Sub\uc9c8\ubb38", ""),
            "Options": row.get("Options", ""),
            "Tags_\ud56d\ubaa9": row.get("Tags_\ud56d\ubaa9", ""),
            "Tags_\ub0b4\uc6a9": row.get("Tags_\ub0b4\uc6a9", ""),
        })
    df_struct = pd.DataFrame(struct_rows, columns=sq_columns) if struct_rows else pd.DataFrame(columns=sq_columns)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_struct.to_excel(writer, sheet_name="\uc9c8\ubb38_\uad6c\uc870\ud654", index=False)

        try:
            if HAS_OPENPYXL_STYLES and struct_rows:
                _apply_structured_styles(writer, struct_rows)
        except Exception as style_err:
            logger.warning("Excel style application failed: %s", str(style_err))

    q_count = sum(1 for r in flat_rows if r.get("\ud589_\uc720\ud615") == "\uc9c8\ubb38")
    stats = {
        "text_rows": text_row_count,
        "table_count": table_count,
        "table_rows": total_table_data_rows,
        "structured_questions": q_count,
    }
    logger.info(
        "Excel saved: %s (tables=%d, structured_questions=%d)",
        output_path, stats["table_count"], stats["structured_questions"],
    )
    return stats


def _apply_excel_styles(
    writer, text_row_count: int, table_excel_rows: List[List[str]]
):
    if not HAS_OPENPYXL_STYLES:
        return

    wb = writer.book
    green_fill = PatternFill(start_color="276749", end_color="276749", fill_type="solid")
    white_bold = Font(bold=True, color="FFFFFF")
    light_green = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    light_orange = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    light_blue = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    dark_gray_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="top")
    bold_font = Font(bold=True)

    ws_text = wb["텍스트_내용"]
    for cell in ws_text[1]:
        cell.fill = green_fill
        cell.font = white_bold
    ws_text.column_dimensions["A"].width = 8
    ws_text.column_dimensions["B"].width = 12
    ws_text.column_dimensions["C"].width = 14
    ws_text.column_dimensions["D"].width = 80

    for row_idx in range(2, text_row_count + 2):
        for cell in ws_text[row_idx]:
            cell.alignment = wrap_align
            cell.border = thin_border
        category_cell = ws_text.cell(row=row_idx, column=2)
        cat_val = str(category_cell.value or "")
        if cat_val == "질문/섹션":
            for cell in ws_text[row_idx]:
                cell.fill = light_green
                cell.font = bold_font
        elif cat_val == "질문ID":
            for cell in ws_text[row_idx]:
                cell.fill = light_orange
                cell.font = bold_font

    ws_table = wb["테이블_원본"]
    for cell in ws_table[1]:
        cell.fill = green_fill
        cell.font = white_bold

    col_widths: Dict[int, int] = {}
    for row_idx in range(2, len(table_excel_rows) + 2):
        row_cells = ws_table[row_idx]
        color_val = str(ws_table.cell(row=row_idx, column=1).value or "")
        col1_val = str(ws_table.cell(row=row_idx, column=2).value or "")

        if col1_val.startswith("[ 테이블"):
            for cell in row_cells:
                cell.fill = light_blue
                cell.font = bold_font
        elif color_val == CC_TITLE:
            for cell in row_cells:
                cell.fill = dark_gray_fill
                cell.font = bold_font
        elif color_val == CC_HEADER:
            for cell in row_cells:
                cell.fill = light_gray_fill

        for cell in row_cells:
            col_idx = cell.column
            val_len = len(str(cell.value or ""))
            if val_len > col_widths.get(col_idx, 0):
                col_widths[col_idx] = val_len

    for col_idx, width in col_widths.items():
        col_letter = ws_table.cell(row=1, column=col_idx).column_letter
        ws_table.column_dimensions[col_letter].width = min(width + 2, 52)


def _apply_structured_styles(writer, struct_rows: list):
    """Apply color coding to 질문_구조화 sheet by row type."""
    if not HAS_OPENPYXL_STYLES:
        return
    sheet_name = "질문_구조화"
    if sheet_name not in writer.sheets:
        return
    ws = writer.sheets[sheet_name]

    green_fill = PatternFill(start_color="276749", end_color="276749", fill_type="solid")
    white_bold = Font(bold=True, color="FFFFFF")
    wrap_align = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    q_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    q_font = Font(bold=True, color="FFFFFF")
    num_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    tag_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

    for cell in ws[1]:
        cell.fill = green_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    widths = [10, 12, 60, 6, 6, 35, 50, 25, 35]
    for i, w in enumerate(widths):
        col_letter = ws.cell(row=1, column=i + 1).column_letter
        ws.column_dimensions[col_letter].width = w

    for row_idx, row_data in enumerate(struct_rows, start=2):
        row_type = row_data.get("행_유형", "")
        for cell in ws[row_idx]:
            cell.alignment = wrap_align
            cell.border = thin_border

        if row_type == "질문":
            for cell in ws[row_idx]:
                cell.fill = q_fill
                cell.font = q_font
        elif row_type == "번호행":
            for cell in ws[row_idx]:
                cell.fill = num_fill
        elif row_type == "태그행":
            for cell in ws[row_idx]:
                cell.fill = tag_fill


# ===========================================================================
# Extract question items for DB compat
# ===========================================================================
def _extract_questions_from_elements(elements: List[Dict]) -> List[QuestionItem]:
    questions: List[QuestionItem] = []
    seen_ids: set = set()

    for elem in elements:
        if elem["type"] != "text":
            continue
        text = elem["data"]
        page_num = elem["page"]

        for line in text.split("\n"):
            line = line.strip()
            if not line:
                continue

            m_sec = SECTION_PATTERN.match(line)
            if m_sec:
                qid = m_sec.group(1)
                if qid not in seen_ids:
                    seen_ids.add(qid)
                    content = m_sec.group(2).strip()
                    points = None
                    pm = POINTS_PATTERN.search(content)
                    if pm:
                        try:
                            points = float(pm.group(1))
                        except (ValueError, TypeError):
                            pass
                    questions.append(QuestionItem(
                        question_id=qid,
                        question_text=content,
                        page_num=page_num,
                        points=points,
                    ))
                continue

            m_qid = QUESTION_ID_PATTERN.match(line)
            if m_qid:
                qid = m_qid.group(1)
                if qid not in seen_ids:
                    seen_ids.add(qid)
                    content = m_qid.group(2).strip()
                    points = None
                    pm = POINTS_PATTERN.search(content)
                    if pm:
                        try:
                            points = float(pm.group(1))
                        except (ValueError, TypeError):
                            pass
                    questions.append(QuestionItem(
                        question_id=qid,
                        question_text=content,
                        page_num=page_num,
                        points=points,
                    ))

    return questions


# ===========================================================================
# Validation
# ===========================================================================
def _validate_extraction(
    text_rows: int,
    table_count: int,
    total_pages: int,
) -> ValidationResult:
    score = 1.0
    messages: List[str] = []

    if text_rows == 0:
        score -= 0.3
        messages.append("텍스트가 추출되지 않았습니다")

    if table_count == 0:
        score -= 0.2
        messages.append("테이블이 추출되지 않았습니다")

    if text_rows > 0 and text_rows < total_pages:
        score -= 0.1
        messages.append(
            "텍스트 행(%d)이 전체 페이지(%d)보다 적습니다" % (text_rows, total_pages)
        )

    score = max(0.0, score)

    if text_rows == 0 and table_count == 0:
        return ValidationResult(
            is_valid=False,
            confidence_score=round(score, 2),
            errors=["텍스트와 테이블 모두 추출되지 않았습니다"],
            warnings=messages,
            needs_human_review=True,
        )

    return ValidationResult(
        is_valid=True,
        confidence_score=round(score, 2),
        errors=[],
        warnings=messages,
        needs_human_review=score < 0.5,
    )


# ===========================================================================
# Document Type Classification
# ===========================================================================
PDF_TYPE_A = "question_guide"       # Question Guide (Reporting Guidance)
PDF_TYPE_B = "scoring_methodology"  # Scoring Methodology

PDF_TYPE_LABELS = {
    PDF_TYPE_A: "질문 가이드 (Question Guide)",
    PDF_TYPE_B: "채점 방법론 (Scoring Methodology)",
}

# Keywords for classification (case-insensitive)
_TYPE_A_KEYWORDS = ["reporting guidance", "questionnaire"]
_TYPE_B_KEYWORDS = ["scoring methodology", "point allocation", "scoring criteria"]


def classify_pdf_type(pdf_path: str, max_pages: int = 5) -> Dict[str, Any]:
    """
    Analyze first N pages of a PDF to classify document type.

    Returns:
        {
            "pdf_type": "question_guide" | "scoring_methodology",
            "confidence": float (0.0~1.0),
            "matched_keywords": list[str],
            "label": str,  # Korean display label
            "ambiguous": bool,  # True if both types detected or neither
        }
    """
    if not HAS_PDFPLUMBER:
        raise ImportError("pdfplumber is required for PDF classification")

    # Priority 1: Filename-based detection
    fname_lower = Path(pdf_path).name.lower()
    if "scoring" in fname_lower:
        return {
            "pdf_type": PDF_TYPE_B,
            "confidence": 1.0,
            "matched_keywords": ["filename:" + Path(pdf_path).name],
            "label": PDF_TYPE_LABELS[PDF_TYPE_B],
            "ambiguous": False,
        }

    text_combined = ""
    with pdfplumber.open(pdf_path) as pdf:
        pages_to_scan = min(max_pages, len(pdf.pages))
        for i in range(pages_to_scan):
            page_text = pdf.pages[i].extract_text() or ""
            text_combined += page_text + "\n"

    text_lower = text_combined.lower()

    a_matches = [kw for kw in _TYPE_A_KEYWORDS if kw in text_lower]
    b_matches = [kw for kw in _TYPE_B_KEYWORDS if kw in text_lower]

    a_score = len(a_matches)
    b_score = len(b_matches)

    if a_score > 0 and b_score == 0:
        return {
            "pdf_type": PDF_TYPE_A,
            "confidence": min(1.0, 0.5 + a_score * 0.25),
            "matched_keywords": a_matches,
            "label": PDF_TYPE_LABELS[PDF_TYPE_A],
            "ambiguous": False,
        }
    elif b_score > 0 and a_score == 0:
        return {
            "pdf_type": PDF_TYPE_B,
            "confidence": min(1.0, 0.5 + b_score * 0.25),
            "matched_keywords": b_matches,
            "label": PDF_TYPE_LABELS[PDF_TYPE_B],
            "ambiguous": False,
        }
    elif a_score > 0 and b_score > 0:
        # Both detected — pick higher score, mark ambiguous
        if a_score >= b_score:
            pdf_type = PDF_TYPE_A
        else:
            pdf_type = PDF_TYPE_B
        return {
            "pdf_type": pdf_type,
            "confidence": 0.4,
            "matched_keywords": a_matches + b_matches,
            "label": PDF_TYPE_LABELS[pdf_type],
            "ambiguous": True,
        }
    else:
        # No keywords found — default to question guide, mark ambiguous
        return {
            "pdf_type": PDF_TYPE_A,
            "confidence": 0.2,
            "matched_keywords": [],
            "label": PDF_TYPE_LABELS[PDF_TYPE_A],
            "ambiguous": True,
        }


# ===========================================================================
# Scoring Methodology Parser (Type B)
# ===========================================================================
# Pattern: "1.5 - Scoring criteria" or "C1.1 Scoring criteria"
SCORING_HEADER_PATTERN = re.compile(
    r"^(\d+(?:\.\d+[a-z]?)*)\s*[-–]\s*[Ss]coring\s+criteria"
)
# Pattern: "(1.5) - Provide details..." — question reference in scoring docs
SCORING_QUESTION_REF_PATTERN = re.compile(
    r"^\((\d+(?:\.\d+[a-z]?)*)\)\s*[-–]?\s*(.*)"
)
# Pattern: point values like "2/2 points", "1/1 point", "0/0 points", "A maximum of X/Y"
POINTS_MAX_PATTERN = re.compile(
    r"[Aa]\s+maximum\s+of\s+(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)\s+points?"
)
POINTS_SIMPLE_PATTERN = re.compile(
    r"(\d+(?:\.\d+)?)\s+[Pp]oints?"
)
# D/A/M/L criteria headers
CRITERIA_HEADER_PATTERN = re.compile(
    r"^(Disclosure|Awareness|Management|Leadership)\s+criteria$",
    re.IGNORECASE,
)
# Route pattern: "ROUTE A)" or "ROUTE B)"
ROUTE_PATTERN = re.compile(r"^(ROUTE\s+[A-Z])\)", re.IGNORECASE)


def _extract_scoring_elements(
    pdf_path: str,
    page_start: Optional[int] = None,
    page_end: Optional[int] = None,
) -> Tuple[List[Dict], int]:
    """Extract elements from a Scoring Methodology PDF."""
    if not HAS_PDFPLUMBER:
        raise ImportError("pdfplumber is required for PDF parsing")

    elements: List[Dict] = []
    total_pages = 0

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        start_idx = (page_start - 1) if page_start else 0
        end_idx = page_end if page_end else total_pages
        start_idx = max(0, start_idx)
        end_idx = min(total_pages, end_idx)

        for page_idx in range(start_idx, end_idx):
            page = pdf.pages[page_idx]
            page_num = page_idx + 1

            try:
                page_rects = _get_page_rects(page)

                # Extract tables with color classification
                tables = page.find_tables()
                table_bboxes = []
                for tbl in tables:
                    try:
                        bbox = tbl.bbox
                        table_bboxes.append(bbox)
                        table_data = tbl.extract()
                        classified_rows = _classify_table_by_color(
                            table_data, bbox, page_rects
                        )
                        elements.append({
                            "type": "table",
                            "page": page_num,
                            "y_top": bbox[1],
                            "data": table_data,
                            "bbox": bbox,
                            "classified_rows": classified_rows,
                        })
                    except Exception as te:
                        logger.warning("Page %d: scoring table error: %s", page_num, str(te))

                # Extract text outside tables — split by question refs for y-ordering
                try:
                    def not_in_table(obj, _bboxes=table_bboxes):
                        obj_y_mid = (obj.get("top", 0) + obj.get("bottom", 0)) / 2.0
                        for bbox in _bboxes:
                            if (bbox[1] - 2) <= obj_y_mid <= (bbox[3] + 2):
                                return False
                        return True
                    filtered_page = page.filter(not_in_table)
                    words = filtered_page.extract_words() or []
                    text_outside = filtered_page.extract_text() or ""
                except Exception:
                    words = []
                    text_outside = page.extract_text() or ""

                if text_outside.strip():
                    # Build word position map for question refs
                    qref_positions = {}  # "(X.Y)" -> y_top
                    for w in words:
                        wt = w.get("text", "")
                        if wt.startswith("(") and ")" in wt:
                            qref_positions[wt] = w.get("top", 0)

                    # Split text at question refs to assign y positions
                    lines = text_outside.split("\n")
                    current_segment = []
                    current_y = 0.0

                    for line in lines:
                        ls = line.strip()
                        if not ls:
                            continue
                        # Check if this line starts a new question ref
                        m = SCORING_QUESTION_REF_PATTERN.match(ls)
                        if m:
                            # Flush previous segment
                            if current_segment:
                                elements.append({
                                    "type": "text",
                                    "page": page_num,
                                    "y_top": current_y,
                                    "data": "\n".join(current_segment),
                                })
                                current_segment = []
                            # Find y position of this question ref
                            qkey = "(%s)" % m.group(1)
                            current_y = qref_positions.get(qkey, current_y + 0.1)
                        current_segment.append(ls)

                    if current_segment:
                        elements.append({
                            "type": "text",
                            "page": page_num,
                            "y_top": current_y,
                            "data": "\n".join(current_segment),
                        })
            except Exception as page_err:
                logger.error("Page %d scoring parse error: %s", page_num, str(page_err))

    elements.sort(key=lambda e: (e["page"], e["y_top"]))
    return elements, total_pages


def _detect_point_allocation_table(table_data: List[List]) -> bool:
    """Check if a table is a Point Allocation table (header, data, or header-only)."""
    if not table_data:
        return False
    all_text = ""
    for row in table_data:
        all_text += " ".join(_clean_cell(c).lower() for c in (row or [])) + " "
    return ("numerator" in all_text or "denominator" in all_text) and (
        "disclosure" in all_text or "awareness" in all_text
    )


def _is_pa_data_only_table(table_data: List[List]) -> bool:
    """Check if a table is a PA data-only row (no headers, just numbers).
    e.g. [['2', '2', '1', '1', '0 or 1', '0 or 1', '0', '0']]"""
    if not table_data or len(table_data) != 1:
        return False
    row = table_data[0]
    non_empty = [_clean_cell(c) for c in (row or []) if _clean_cell(c)]
    if len(non_empty) < 4:
        return False
    # All non-empty values should be numeric-looking
    return all(v[0].isdigit() or v.startswith("0") for v in non_empty)


def _extract_point_allocation(table_data: List[List]) -> Dict[str, str]:
    """Extract Point Allocation table into {D_num, D_den, A_num, ...} dict.

    Real PDF structure (24 columns, many empty):
      Row 0: ['', 'Disclosure', '', '', 'Disclosure', '', '', 'Awareness', ...]
      Row 1: ['', 'numerator',  '', '', 'denominator','', '', 'numerator', ...]
      Row 2: ['', '7.5',        '', '', '7.5',        '', '', '4',         ...]

    Strategy: strip empty cells from each row → pair headers with data positionally.
    """
    if not table_data or len(table_data) < 2:
        return {}

    # Strip empty cells from each row, keeping only non-empty values
    def _strip_empty(row):
        return [_clean_cell(c) for c in (row or []) if _clean_cell(c)]

    # Find header rows (contain level names) and data row (contains numbers)
    level_row = []   # e.g. ['Disclosure', 'Disclosure', 'Awareness', ...]
    suffix_row = []  # e.g. ['numerator', 'denominator', 'numerator', ...]
    data_row = []    # e.g. ['7.5', '7.5', '4', '4', '1 or 2', '2', ...]

    for row in table_data:
        stripped = _strip_empty(row)
        if not stripped:
            continue
        joined = " ".join(stripped).lower()
        if "disclosure" in joined and "awareness" in joined:
            level_row = stripped
        elif "numerator" in joined and "denominator" in joined:
            suffix_row = stripped
        elif stripped and (stripped[0][0:1].isdigit() or stripped[0].startswith("0")):
            data_row = stripped

    if not data_row:
        return {}

    # Build 8-column mapping: pair level+suffix → data value
    # Expected order: D_num, D_den, A_num, A_den, M_num, M_den, L_num, L_den
    level_map = {"disclosure": "D", "awareness": "A",
                 "management": "M", "leadership": "L"}

    result = {}
    for i, val in enumerate(data_row):
        # Determine level from level_row
        level_short = ""
        if i < len(level_row):
            for full, short in level_map.items():
                if full in level_row[i].lower():
                    level_short = short
                    break

        # Determine suffix from suffix_row
        suffix = ""
        if i < len(suffix_row):
            sl = suffix_row[i].lower()
            if "numerator" in sl:
                suffix = "num"
            elif "denominator" in sl:
                suffix = "den"

        if level_short and suffix:
            key = level_short + "_" + suffix
            result[key] = val

    return result


def _build_scoring_questions(elements: List[Dict]) -> List[Dict]:
    """
    Build ONE ROW per question from CDP Scoring Methodology PDF.

    Each row contains the full D/A/M/L criteria + point allocation in columns,
    clearly distinct from the questionnaire parser output.

    Output columns per row:
      문항ID, 질문내용, 페이지, 최대배점,
      Disclosure_기준, Disclosure_배점,
      Awareness_기준, Awareness_배점,
      Management_기준, Management_배점,
      Leadership_기준, Leadership_배점,
      D_num, D_den, A_num, A_den, M_num, M_den, L_num, L_den,
      테마, 섹터
    """
    questions: List[Dict] = []  # final output
    scoring_started = False  # True after "X.Y - Scoring criteria" header
    pending_pa_header = None  # Header-only PA table waiting for data row

    # -- Current question accumulator --
    cur: Optional[Dict] = None
    current_level = ""       # "Disclosure" | "Awareness" | ...
    level_text_buf: List[str] = []

    def _flush_level():
        nonlocal level_text_buf, current_level
        if cur and current_level and level_text_buf:
            key = current_level + "_기준"
            existing = cur.get(key, "")
            text = " ".join(level_text_buf).strip()
            cur[key] = (existing + " " + text).strip() if existing else text
        level_text_buf = []

    def _flush_question():
        nonlocal cur, current_level
        _flush_level()
        if cur and cur.get("문항ID"):
            questions.append(cur)
        cur = None
        current_level = ""

    def _new_question(qid: str, page: int):
        nonlocal cur, current_level
        _flush_question()
        cur = {
            "문항ID": qid,
            "질문내용": "",
            "페이지": page,
            "최대배점": "",
            "Disclosure_기준": "", "Disclosure_배점": "",
            "Awareness_기준": "", "Awareness_배점": "",
            "Management_기준": "", "Management_배점": "",
            "Leadership_기준": "", "Leadership_배점": "",
            "D_num": "", "D_den": "", "A_num": "", "A_den": "",
            "M_num": "", "M_den": "", "L_num": "", "L_den": "",
            "테마": "", "섹터": "",
        }
        current_level = ""

    for elem in elements:
        if elem["type"] == "text":
            for line in elem["data"].split("\n"):
                ls = line.strip()
                if not ls:
                    continue
                # Skip noise
                if ls.startswith("Page ") and " out of " in ls:
                    continue
                if ls.startswith("@cdp") or ls == "CDP" or ls.startswith("\u25d1"):
                    continue

                # 1) Question reference: "(1.5) - Provide details..."
                m = SCORING_QUESTION_REF_PATTERN.match(ls)
                if m:
                    qid = m.group(1)
                    qtxt = m.group(2).strip()
                    _new_question(qid, elem["page"])
                    cur["질문내용"] = qtxt
                    scoring_started = False  # wait for "X.Y - Scoring criteria"
                    continue

                # 2) Scoring section header: "1.5 - Scoring criteria"
                m = SCORING_HEADER_PATTERN.match(ls)
                if m:
                    hdr_id = m.group(1)
                    if cur is None:
                        _new_question(hdr_id, elem["page"])
                    elif cur.get("문항ID") != hdr_id:
                        _new_question(hdr_id, elem["page"])
                    scoring_started = True
                    continue

                # No current question context → skip
                if cur is None:
                    continue

                # 3) D/A/M/L criteria header
                m = CRITERIA_HEADER_PATTERN.match(ls)
                if m:
                    _flush_level()
                    current_level = m.group(1).capitalize()
                    continue

                # 4) Route: "ROUTE A)" / "ROUTE B)"
                m = ROUTE_PATTERN.match(ls)
                if m:
                    _flush_level()
                    route_label = m.group(1).upper()
                    rest = ls[m.end():].strip()
                    level_text_buf.append("[%s] %s" % (route_label, rest))
                    continue

                # 5) "Not scored."
                if ls.lower() == "not scored.":
                    _flush_level()
                    if current_level:
                        cur[current_level + "_기준"] = "Not scored."
                        cur[current_level + "_배점"] = "0"
                    continue

                # 6) Max points: "A maximum of X/Y points..."
                m = POINTS_MAX_PATTERN.search(ls)
                if m:
                    numerator = m.group(1)
                    denominator = m.group(2)
                    pts_str = "%s/%s" % (numerator, denominator)
                    is_question_level = "for this question" in ls.lower()

                    if not current_level:
                        # Outside D/A/M/L section → overall max
                        cur["최대배점"] = pts_str
                    else:
                        # Inside D/A/M/L section
                        _flush_level()
                        cur[current_level + "_배점"] = pts_str
                        # If overall max not set yet and "for this question",
                        # this is the overall question max (appears before Awareness)
                        if not cur["최대배점"] and is_question_level:
                            cur["최대배점"] = pts_str
                    continue

                # 7) Simple points: "X point(s)"
                m = POINTS_SIMPLE_PATTERN.search(ls)
                if m and current_level:
                    level_text_buf.append(ls)
                    _flush_level()
                    pts = m.group(1)
                    cur[current_level + "_배점"] = pts
                    continue

                # 8) "Point Allocation" text header → skip
                if "point allocation" in ls.lower():
                    continue

                # 9) "OR" separator between routes
                if ls == "OR":
                    _flush_level()
                    level_text_buf.append("[OR]")
                    continue

                # 10) Accumulate criteria text
                if current_level:
                    level_text_buf.append(ls)
                # else: general description text, skip for now

        elif elem["type"] == "table":
            _flush_level()
            table_data = elem.get("data") or []

            if _detect_point_allocation_table(table_data):
                alloc = _extract_point_allocation(table_data)
                if alloc:
                    if cur and scoring_started:
                        target = cur
                    elif questions:
                        target = questions[-1]
                    else:
                        target = cur
                    if target:
                        # Only assign if target doesn't already have D_num,
                        # or overwrite with non-zero values (last wins for same question)
                        if not target.get("D_num") or any(v != "0" for v in alloc.values()):
                            for k, v in alloc.items():
                                target[k] = v
                else:
                    pending_pa_header = table_data

            elif _is_pa_data_only_table(table_data):
                # Data-only row from a page-split PA table
                # Build complete table by combining with pending header
                if pending_pa_header:
                    merged = pending_pa_header + table_data
                    pending_pa_header = None
                else:
                    # No pending header — build with standard D/A/M/L order
                    merged = [
                        ["Disclosure", "Disclosure", "Awareness", "Awareness",
                         "Management", "Management", "Leadership", "Leadership"],
                        ["numerator", "denominator", "numerator", "denominator",
                         "numerator", "denominator", "numerator", "denominator"],
                    ] + table_data

                alloc = _extract_point_allocation(merged)
                if alloc:
                    # Split PA data always belongs to the PREVIOUS question
                    # (cur just received a new question ref before this data arrived)
                    if questions:
                        target = questions[-1]
                    else:
                        target = cur
                    if target:
                        if not target.get("D_num") or any(v != "0" for v in alloc.values()):
                            for k, v in alloc.items():
                                target[k] = v

            else:
                # Theme/Sector table
                target = cur
                if target and table_data:
                    all_vals = []
                    for row in table_data:
                        for c in (row or []):
                            v = _clean_cell(c)
                            if v and v.lower() not in ("theme", "sector that scoring criteria apply to",
                                                        "sector that scoring criteria a"):
                                if v not in all_vals:
                                    all_vals.append(v)
                    if all_vals:
                        target["테마"] = all_vals[0]
                        if len(all_vals) > 1:
                            target["섹터"] = " ".join(all_vals[1:])

    _flush_question()


    return questions


def _extract_pa_by_y_interleave(pdf_path: str, page_start=None, page_end=None) -> Dict[str, Dict]:
    """Extract Point Allocation values by processing text and tables in y-coordinate order.
    Returns {qid: {D_num, D_den, A_num, A_den, M_num, M_den, L_num, L_den}}."""
    if not HAS_PDFPLUMBER:
        return {}

    def _clean(c):
        if c is None:
            return ""
        s = str(c).strip()
        return "" if s.lower() in ("nan", "none") else s

    def _extract_pa(table_data):
        rows = [[_clean(c) for c in (row or []) if _clean(c)] for row in table_data]
        rows = [r for r in rows if r]
        if not rows:
            return None
        level_row = suffix_row = data_row = None
        for r in rows:
            j = " ".join(r).lower()
            if "disclosure" in j and "awareness" in j:
                level_row = r
            elif "numerator" in j and "denominator" in j:
                suffix_row = r
            elif r and (r[0][0:1].isdigit() or r[0].startswith("0")):
                data_row = r
        if not data_row:
            return None
        if not level_row:
            level_row = ["Disclosure"] * 2 + ["Awareness"] * 2 + ["Management"] * 2 + ["Leadership"] * 2
        if not suffix_row:
            suffix_row = ["numerator", "denominator"] * 4
        lm = {"disclosure": "D", "awareness": "A", "management": "M", "leadership": "L"}
        result = {}
        for i, val in enumerate(data_row):
            lv = ""
            if i < len(level_row):
                for f, s in lm.items():
                    if f in level_row[i].lower():
                        lv = s
                        break
            sx = ""
            if i < len(suffix_row):
                sl = suffix_row[i].lower()
                if "numerator" in sl:
                    sx = "num"
                elif "denominator" in sl:
                    sx = "den"
            if lv and sx:
                result[lv + "_" + sx] = val
        return result if len(result) >= 4 else None

    qref_pat = re.compile(r"\((\d+(?:\.\d+[a-z]?)*)\)\s*[-\u2013]")
    ground_truth = {}
    last_qid = None
    pending_header = None

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        si = (page_start - 1) if page_start else 0
        ei = page_end if page_end else total

        for pi in range(max(0, si), min(total, ei)):
            page = pdf.pages[pi]
            events = []

            try:
                words = page.extract_words() or []
                text = page.extract_text() or ""
                for m in qref_pat.finditer(text):
                    qid = m.group(1)
                    target = "(" + qid + ")"
                    y = 0
                    for w in words:
                        if target in w.get("text", ""):
                            y = w.get("top", 0)
                            break
                    else:
                        y = (m.start() / max(len(text), 1)) * page.height
                    events.append((y, "qref", qid))
            except Exception:
                pass

            for tbl in page.find_tables():
                try:
                    data = tbl.extract()
                    if data:
                        events.append((tbl.bbox[1], "table", data))
                except Exception:
                    pass

            events.sort(key=lambda x: x[0])

            for _, etype, edata in events:
                if etype == "qref":
                    last_qid = edata
                elif etype == "table":
                    at = " ".join(_clean(c).lower() for row in edata for c in (row or []))
                    is_pa = ("numerator" in at or "denominator" in at) and "disclosure" in at
                    stripped = [[_clean(c) for c in (row or []) if _clean(c)] for row in edata]
                    stripped = [r for r in stripped if r]
                    is_data_only = (len(stripped) == 1 and len(stripped[0]) >= 4 and
                                    all(v[0:1].isdigit() or v.startswith("0") for v in stripped[0]))

                    if is_pa:
                        pa = _extract_pa(edata)
                        if pa and last_qid:
                            ground_truth[last_qid] = pa
                            pending_header = None
                        elif not pa:
                            pending_header = edata
                    elif is_data_only:
                        if pending_header:
                            merged = pending_header + edata
                            pa = _extract_pa(merged)
                            if pa and last_qid:
                                ground_truth[last_qid] = pa
                            pending_header = None
                        else:
                            std_header = [
                                ["Disclosure", "Disclosure", "Awareness", "Awareness",
                                 "Management", "Management", "Leadership", "Leadership"],
                                ["numerator", "denominator", "numerator", "denominator",
                                 "numerator", "denominator", "numerator", "denominator"],
                            ]
                            pa = _extract_pa(std_header + edata)
                            if pa and last_qid:
                                ground_truth[last_qid] = pa

    return ground_truth


def _save_scoring_excel(
    elements: List[Dict],
    output_path: str,
    source_file: str = "",
    pdf_path: str = "",
    page_start: Optional[int] = None,
    page_end: Optional[int] = None,
) -> Dict[str, int]:
    """Save scoring methodology data as 1 row per question with D/A/M/L columns."""
    if not HAS_PANDAS:
        raise ImportError("pandas is required for Excel output")

    scoring_questions = _build_scoring_questions(elements)

    # Override PA values from accurate y-interleaved extraction
    if pdf_path:
        pa_truth = _extract_pa_by_y_interleave(pdf_path, page_start, page_end)
        for q in scoring_questions:
            qid = q.get("문항ID", "")
            if qid in pa_truth:
                for k, v in pa_truth[qid].items():
                    q[k] = v

    columns = [
        "문항ID", "질문내용", "페이지", "최대배점",
        "Disclosure_기준", "Disclosure_배점",
        "Awareness_기준", "Awareness_배점",
        "Management_기준", "Management_배점",
        "Leadership_기준", "Leadership_배점",
        "D_num", "D_den", "A_num", "A_den",
        "M_num", "M_den", "L_num", "L_den",
        "테마", "섹터",
    ]
    struct_rows = []
    for q in scoring_questions:
        struct_rows.append({col: q.get(col, "") for col in columns})

    df = pd.DataFrame(struct_rows, columns=columns) if struct_rows else pd.DataFrame(columns=columns)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="채점_방법론", index=False)

        if HAS_OPENPYXL_STYLES and struct_rows:
            try:
                wb = writer.book
                ws = wb["채점_방법론"]
                DARK_GREEN = PatternFill("solid", fgColor="1F5C35")
                LIGHT_GREEN = PatternFill("solid", fgColor="D9EAD3")
                WHITE = PatternFill("solid", fgColor="FFFFFF")
                HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
                WRAP = Alignment(wrap_text=True, vertical="top")

                # Header row
                for cell in ws[1]:
                    cell.fill = DARK_GREEN
                    cell.font = HDR_FONT
                    cell.alignment = WRAP

                # Data rows — alternating green/white
                for ri in range(2, len(struct_rows) + 2):
                    fill = LIGHT_GREEN if (ri % 2 == 0) else WHITE
                    for cell in ws[ri]:
                        cell.fill = fill
                        cell.font = Font(name="Calibri", size=10)
                        cell.alignment = WRAP

                # Auto-fit column widths (rough)
                for col_cells in ws.columns:
                    max_len = max(len(str(c.value or "")) for c in col_cells)
                    adjusted = min(max_len + 2, 50)
                    ws.column_dimensions[col_cells[0].column_letter].width = adjusted
            except Exception as e:
                logger.warning("Scoring style error: %s", e)

    q_count = len(scoring_questions)
    stats = {
        "text_rows": sum(1 for e in elements if e["type"] == "text"),
        "table_count": sum(1 for e in elements if e["type"] == "table"),
        "table_rows": sum(len(e.get("data") or []) for e in elements if e["type"] == "table"),
        "structured_questions": q_count,
    }
    logger.info("Scoring Excel saved: %s (questions=%d)", output_path, q_count)
    return stats


def run_scoring_parser(
    pdf_path: str,
    output_dir: Optional[str] = None,
    save_excel: bool = True,
    page_start: Optional[int] = None,
    page_end: Optional[int] = None,
) -> AgentResult:
    """Parse a Scoring Methodology PDF and output structured Excel."""
    print(f"\n[run_scoring_parser] ENTERED - file={Path(pdf_path).name}")
    start_time = time.time()
    warnings_list: List[str] = []
    pdf_file = Path(pdf_path)

    if not pdf_file.exists():
        return AgentResult(
            agent_name="ScoringMethodologyParser",
            status=AgentStatus.FAILED,
            error_message="PDF 파일을 찾을 수 없습니다: %s" % pdf_path,
            data={},
            processing_time_sec=round(time.time() - start_time, 2),
        )

    try:
        elements, total_pages = _extract_scoring_elements(
            pdf_path, page_start=page_start, page_end=page_end
        )
    except Exception as ex:
        return AgentResult(
            agent_name="ScoringMethodologyParser",
            status=AgentStatus.FAILED,
            error_message="PDF 추출 실패: %s" % str(ex),
            data={},
            processing_time_sec=round(time.time() - start_time, 2),
        )

    excel_path = None
    excel_filename = None
    excel_stats = {"text_rows": 0, "table_count": 0, "table_rows": 0, "structured_questions": 0}

    if save_excel:
        stem = pdf_file.stem
        if page_start or page_end:
            ps = page_start or 1
            pe = page_end or total_pages
            suffix = "_p%d-%d_scoring.xlsx" % (ps, pe)
        else:
            suffix = "_scoring.xlsx"
        excel_filename = stem + suffix
        print(f"[run_scoring_parser] excel_filename={excel_filename}")

        if output_dir:
            out_dir = Path(output_dir)
        else:
            out_dir = Path("c:/Project/CDP-AI-Platform/data/outputs")
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_path = str(out_dir / excel_filename)

        try:
            excel_stats = _save_scoring_excel(
                elements, excel_path, source_file=pdf_file.name,
                pdf_path=pdf_path, page_start=page_start, page_end=page_end,
            )
            print(f"[run_scoring_parser] Excel SAVED: {excel_path}")
            print(f"[run_scoring_parser] stats={excel_stats}")
        except Exception as ex:
            msg = "Excel 저장 실패: %s" % str(ex)
            print(f"[run_scoring_parser] Excel SAVE FAILED: {ex}")
            logger.error(msg)
            warnings_list.append(msg)
            excel_path = None

    validation = _validate_extraction(
        excel_stats["text_rows"], excel_stats["table_count"], total_pages
    )
    warnings_list.extend(validation.warnings)

    elapsed = time.time() - start_time
    page_range_str = None
    if page_start or page_end:
        ps = page_start or 1
        pe = page_end or total_pages
        page_range_str = "%d-%d" % (ps, pe)

    data = {
        "source_file": pdf_file.name,
        "pdf_type": PDF_TYPE_B,
        "pdf_type_label": PDF_TYPE_LABELS[PDF_TYPE_B],
        "total_questions": excel_stats["structured_questions"],
        "page_range": page_range_str,
        "excel_path": excel_path,
        "excel_filename": excel_filename,
        "excel_stats": excel_stats,
        "parse_warnings": warnings_list,
        "tables_extracted": excel_stats["table_count"],
    }

    if not validation.is_valid:
        return AgentResult(
            agent_name="ScoringMethodologyParser",
            status=AgentStatus.FAILED,
            error_message="PDF 파싱 실패: %s" % "; ".join(validation.errors),
            data=data,
            processing_time_sec=elapsed,
            validation=validation,
        )

    return AgentResult(
        agent_name="ScoringMethodologyParser",
        status=AgentStatus.SUCCESS,
        data=data,
        processing_time_sec=elapsed,
        validation=validation,
        created_at=time.strftime("%Y-%m-%dT%H:%M:%S"),
    )


# ===========================================================================
# Main function
# ===========================================================================
def run_pdf_parser(
    pdf_path: str,
    output_dir: Optional[str] = None,
    save_excel: bool = True,
    page_start: Optional[int] = None,
    page_end: Optional[int] = None,
    session_id: Optional[str] = None,
    pdf_type: Optional[str] = None,
) -> AgentResult:
    """
    Main PDF parser entry point.

    Args:
        pdf_type: "question_guide" or "scoring_methodology".
                  If None, auto-classifies via classify_pdf_type().
    """
    start_time = time.time()
    warnings_list: List[str] = []
    pdf_file = Path(pdf_path)

    # ===== 디버그: 무조건 콘솔 출력 =====
    print(f"\n{'='*60}")
    print(f"[run_pdf_parser] CALLED")
    print(f"  file: {pdf_file.name}")
    print(f"  pdf_type param: {pdf_type!r}")
    print(f"{'='*60}")

    if not pdf_file.exists():
        return AgentResult(
            agent_name="PDFParserAgent",
            status=AgentStatus.FAILED,
            error_message="PDF 파일을 찾을 수 없습니다: %s" % pdf_path,
            data={},
            processing_time_sec=round(time.time() - start_time, 2),
        )

    # --- Document type branching (파일명 키워드 기반 — 절대 원칙) ---
    # 파일명에 scoring이 포함되면 무조건 TYPE_B (pdf_type 파라미터와 무관)
    fname_lower = pdf_file.name.lower().replace(" ", "_")
    if "scoring" in fname_lower:
        pdf_type = PDF_TYPE_B
    elif pdf_type is None:
        if "questionnaire" in fname_lower:
            pdf_type = PDF_TYPE_A
        else:
            pdf_type = PDF_TYPE_A

    print(f"[run_pdf_parser] RESOLVED pdf_type={pdf_type!r} for '{pdf_file.name}'")

    if pdf_type == PDF_TYPE_B:
        print(f"[run_pdf_parser] >>> ROUTING TO run_scoring_parser()")
        return run_scoring_parser(
            pdf_path=pdf_path,
            output_dir=output_dir,
            save_excel=save_excel,
            page_start=page_start,
            page_end=page_end,
        )

    logger.info(">>> Routing to questionnaire parser (Type A)")

    db_module = None
    try:
        from backend.core import database as db_module
    except ImportError:
        db_module = None

    try:
        skill = load_skill("pdf_extraction")
    except Exception:
        skill = None

    db_session = None
    if db_module and session_id:
        try:
            db_session = db_module.get_session(session_id)
        except Exception:
            db_session = None

    try:
        elements, total_pages = _extract_page_elements(
            pdf_path, page_start=page_start, page_end=page_end
        )
    except Exception as ex:
        return AgentResult(
            agent_name="PDFParserAgent",
            status=AgentStatus.FAILED,
            error_message="PDF 추출 실패: %s" % str(ex),
            data={},
            processing_time_sec=round(time.time() - start_time, 2),
        )

    questions = _extract_questions_from_elements(elements)

    excel_path = None
    excel_filename = None
    excel_stats = {"text_rows": 0, "table_count": 0, "table_rows": 0, "structured_questions": 0}

    if save_excel:
        stem = pdf_file.stem
        if page_start or page_end:
            ps = page_start or 1
            pe = page_end or total_pages
            suffix = "_p%d-%d_parsed.xlsx" % (ps, pe)
        else:
            suffix = "_parsed.xlsx"
        excel_filename = stem + suffix

        if output_dir:
            out_dir = Path(output_dir)
        else:
            # Default to configured outputs directory
            out_dir = Path("c:/Project/CDP-AI-Platform/data/outputs")
        out_dir.mkdir(parents=True, exist_ok=True)
        excel_path = str(out_dir / excel_filename)

        try:
            excel_stats = _save_faithful_excel(
                elements, excel_path, source_file=pdf_file.name,
            )
        except Exception as ex:
            msg = "Excel 저장 실패: %s" % str(ex)
            logger.error(msg)
            warnings_list.append(msg)
            excel_path = None

    validation = _validate_extraction(
        excel_stats["text_rows"],
        excel_stats["table_count"],
        total_pages,
    )
    warnings_list.extend(validation.warnings)

    if db_session and db_module:
        try:
            parse_result = PDFParseResult(
                source_file=pdf_file.name,
                total_pages=total_pages,
                questions_found=len(questions),
                excel_path=excel_path,
            )
            db_module.save_parse_result(db_session, parse_result)
        except Exception as db_err:
            logger.warning("DB 저장 실패: %s", str(db_err))

    tables_extracted = excel_stats["table_count"]
    elapsed = time.time() - start_time

    page_range_str = None
    if page_start or page_end:
        ps = page_start or 1
        pe = page_end or total_pages
        page_range_str = "%d-%d" % (ps, pe)

    questions_data = [
        {
            "question_id": q.question_id,
            "question_text": q.question_text,
            "page_num": q.page_num,
            "max_points": q.max_points,
        }
        for q in questions[:50]
    ]

    data = {
        "source_file": pdf_file.name,
        "pdf_type": PDF_TYPE_A,
        "pdf_type_label": PDF_TYPE_LABELS[PDF_TYPE_A],
        "total_questions": len(questions),
        "questions": questions_data,
        "page_range": page_range_str,
        "excel_path": excel_path,
        "excel_filename": excel_filename,
        "excel_stats": excel_stats,
        "parse_warnings": warnings_list,
        "tables_extracted": tables_extracted,
    }

    if not validation.is_valid:
        return AgentResult(
            agent_name="PDFParserAgent",
            status=AgentStatus.FAILED,
            error_message="PDF 파싱 실패: %s" % "; ".join(validation.errors),
            data=data,
            processing_time_sec=elapsed,
            validation=validation,
        )

    return AgentResult(
        agent_name="PDFParserAgent",
        status=AgentStatus.SUCCESS,
        data=data,
        processing_time_sec=elapsed,
        validation=validation,
        created_at=time.strftime("%Y-%m-%dT%H:%M:%S"),
    )
