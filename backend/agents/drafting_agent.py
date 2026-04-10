# -*- coding: utf-8 -*-
"""
Drafting Agent — 전년도 CDP 답변 기반 2025년 초안 자동 작성

Flow:
1. 2024 QA 엑셀에서 전년도 답변 로드 (Q_No + Answer)
2. 2025 Master Analysis 에서 문항 목록 로드
3. Q_No Exact Match → 전년도 답변을 초안으로 복사
4. Fuzzy Match → 질문 텍스트 유사도로 매칭
5. AF열 (Answer)에 초안 입력 + 저장
"""
import re
import time
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from difflib import SequenceMatcher

import pandas as pd
from openpyxl import load_workbook

from backend.core.models import AgentResult, AgentStatus
from backend.core.config import settings

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
QA_2024_PATH = "data/knowledge/2024_CDP_QA.xlsx.xlsx"
MASTER_PATH = "data/outputs/CDP_2025_Master_Analysis.xlsx"

# ---------------------------------------------------------------------------
# Progress tracking (global for API polling)
# ---------------------------------------------------------------------------
draft_progress = {
    "running": False, "done": False, "error": "",
    "current": 0, "total": 0, "matched": 0, "fuzzy": 0,
}


def _normalize_qid(raw: str) -> str:
    """'[1.1]' or '(1.1)' → '1.1'"""
    return re.sub(r"[\[\]() ]", "", str(raw).strip())


def _text_similarity(a: str, b: str) -> float:
    """0~1 text similarity using SequenceMatcher."""
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a.lower()[:200], b.lower()[:200]).ratio()


# ---------------------------------------------------------------------------
# Load 2024 Q&A
# ---------------------------------------------------------------------------
def load_2024_qa(base_dir: str = ".") -> Dict[str, Dict]:
    """
    Load 2024 CDP QA data.
    Returns: {qid: {"answer": str, "q_des": str, "subs": [{sub_no, answer}]}}
    """
    path = Path(base_dir) / QA_2024_PATH
    if not path.exists():
        raise FileNotFoundError(f"2024 QA 파일을 찾을 수 없습니다: {path}")

    # Header is at row 2 (0-indexed), data starts row 3
    raw = pd.read_excel(str(path), sheet_name=0, header=None, dtype=str, engine="openpyxl")
    raw = raw.fillna("")

    headers = raw.iloc[2].values
    data = raw.iloc[3:].reset_index(drop=True)

    # Map columns by header names
    col_names = {}
    for i, h in enumerate(headers):
        s = str(h).strip()
        if s and s != "nan":
            col_names[s] = i

    qno_i = col_names.get("Q_No.", 1)
    lv_i = col_names.get("Lv.", 2)
    qdes_i = col_names.get("Q_Des.", 5)
    sub_no_i = col_names.get("1_Sub_Q_No.", None)
    sub_des_i = col_names.get("2_Sub_Q_Des.", None)

    # Answer is the last column (col 31 = "2024 CDP")
    ans_i = data.shape[1] - 1

    # Build Q&A lookup
    qa_lookup: Dict[str, Dict] = {}
    current_qid = ""

    for _, row in data.iterrows():
        qno = _normalize_qid(str(row.iloc[qno_i]))
        lv = str(row.iloc[lv_i]).strip()
        answer = str(row.iloc[ans_i]).strip()
        q_des = str(row.iloc[qdes_i]).strip() if qdes_i < len(row) else ""

        if not qno:
            continue

        if lv == "Lv.1":
            current_qid = qno
            if qno not in qa_lookup:
                qa_lookup[qno] = {"q_des": q_des, "answer": "", "subs": []}
            if answer:
                qa_lookup[qno]["answer"] = answer
            if q_des and not qa_lookup[qno]["q_des"]:
                qa_lookup[qno]["q_des"] = q_des

        elif lv == "Lv.2" and current_qid:
            if current_qid not in qa_lookup:
                qa_lookup[current_qid] = {"q_des": "", "answer": "", "subs": []}

            sub_no = ""
            if sub_no_i is not None and sub_no_i < len(row):
                sub_no = str(row.iloc[sub_no_i]).strip()

            if answer:
                qa_lookup[current_qid]["subs"].append({
                    "sub_no": sub_no,
                    "answer": answer,
                })

    logger.info("2024 QA loaded: %d questions", len(qa_lookup))
    return qa_lookup


# ---------------------------------------------------------------------------
# Main drafting function
# ---------------------------------------------------------------------------
def run_drafting_agent(
    base_dir: str = ".",
    master_path: Optional[str] = None,
    use_fuzzy: bool = True,
    fuzzy_threshold: float = 0.6,
) -> AgentResult:
    """
    2024 답변 기반으로 2025 Master Analysis AF열에 초안을 작성합니다.

    Args:
        use_fuzzy: True면 Q_No 불일치 시 질문 텍스트 유사도로 매칭
        fuzzy_threshold: 유사도 임계값 (0~1)
    """
    global draft_progress
    start_time = time.time()
    warnings: List[str] = []

    draft_progress = {
        "running": True, "done": False, "error": "",
        "current": 0, "total": 0, "matched": 0, "fuzzy": 0,
    }

    m_path = Path(base_dir) / (master_path or MASTER_PATH)
    if not m_path.exists():
        draft_progress["running"] = False
        draft_progress["error"] = "Master 파일 없음"
        return AgentResult(
            agent_name="DraftingAgent",
            status=AgentStatus.FAILED,
            error_message=f"Master 파일을 찾을 수 없습니다: {m_path}. Step 2를 먼저 실행하세요.",
            data={}, processing_time_sec=0,
        )

    # 1. Load 2024 QA
    try:
        qa_2024 = load_2024_qa(base_dir)
    except Exception as e:
        draft_progress["running"] = False
        draft_progress["error"] = str(e)
        return AgentResult(
            agent_name="DraftingAgent",
            status=AgentStatus.FAILED,
            error_message=f"2024 QA 로드 실패: {e}",
            data={}, processing_time_sec=round(time.time() - start_time, 2),
        )

    # 2. Load Master
    master_df = pd.read_excel(str(m_path), dtype=str, engine="openpyxl").fillna("")

    # Ensure Answer column exists
    if "Answer" not in master_df.columns:
        master_df["Answer"] = ""

    lv1_rows = master_df[master_df["Lv"] == "Lv.1"]
    draft_progress["total"] = len(lv1_rows)

    # 3. Build fuzzy index (2024 Q_Des → qid)
    fuzzy_index: List[Tuple[str, str, str]] = []  # (qid, q_des, combined_answer)
    for qid, qa_data in qa_2024.items():
        combined = qa_data["answer"]
        if qa_data["subs"]:
            sub_answers = [s["answer"] for s in qa_data["subs"] if s["answer"]]
            if sub_answers:
                combined = combined + "\n" + "\n".join(sub_answers) if combined else "\n".join(sub_answers)
        if combined.strip():
            fuzzy_index.append((qid, qa_data.get("q_des", ""), combined))

    # 4. Match and draft
    exact_count = 0
    fuzzy_count = 0
    drafted_count = 0

    for idx in master_df.index:
        row = master_df.loc[idx]
        if row["Lv"] != "Lv.1":
            continue

        qid = row["Q_No"]
        q_des = row.get("Q_Des", "")
        draft_progress["current"] += 1

        # Skip if already has answer
        if row.get("Answer", "").strip():
            continue

        draft_text = ""

        # Exact match
        if qid in qa_2024:
            qa = qa_2024[qid]
            combined = qa["answer"]
            if qa["subs"]:
                sub_parts = [f"[{s['sub_no']}] {s['answer']}" if s['sub_no'] else s['answer']
                             for s in qa["subs"] if s["answer"]]
                if sub_parts:
                    combined = combined + "\n" + "\n".join(sub_parts) if combined else "\n".join(sub_parts)
            if combined.strip():
                draft_text = combined.strip()
                exact_count += 1

        # Fuzzy match
        if not draft_text and use_fuzzy and q_des:
            best_sim = 0.0
            best_match = ""
            best_qid = ""
            for fqid, fdes, fans in fuzzy_index:
                sim = _text_similarity(q_des, fdes)
                if sim > best_sim and sim >= fuzzy_threshold:
                    best_sim = sim
                    best_match = fans
                    best_qid = fqid
            if best_match:
                draft_text = f"[Ref: 2024 Q{best_qid}, sim={best_sim:.0%}]\n{best_match}"
                fuzzy_count += 1

        if draft_text:
            master_df.at[idx, "Answer"] = draft_text
            drafted_count += 1

        draft_progress["matched"] = exact_count
        draft_progress["fuzzy"] = fuzzy_count

    # 5. Also fill Lv.2 sub-answers where possible
    current_qid = ""
    for idx in master_df.index:
        row = master_df.loc[idx]
        if row["Lv"] == "Lv.1":
            current_qid = row["Q_No"]
        elif row["Lv"] == "Lv.2" and current_qid:
            if row.get("Answer", "").strip():
                continue
            sub_no = row.get("Sub_No", "")
            qa = qa_2024.get(current_qid, {})
            for s in qa.get("subs", []):
                if s["sub_no"] == sub_no and s["answer"]:
                    master_df.at[idx, "Answer"] = s["answer"]
                    break

    # 6. Save back to Excel
    try:
        wb = load_workbook(str(m_path))
        ws = wb.active

        # Find Answer column index
        ans_col_idx = None
        for ci, cell in enumerate(ws[1], 1):
            if cell.value == "Answer":
                ans_col_idx = ci
                break

        if ans_col_idx is None:
            # Add Answer column
            ans_col_idx = ws.max_column + 1
            ws.cell(1, ans_col_idx, "Answer")

        # Write answers
        for ri, idx in enumerate(master_df.index, 2):
            val = master_df.at[idx, "Answer"]
            if val:
                ws.cell(ri, ans_col_idx, str(val))

        wb.save(str(m_path))
        wb.close()
        logger.info("Drafts saved to %s (col %d)", m_path, ans_col_idx)
    except Exception as e:
        warnings.append(f"Excel 저장 실패: {e}")

    elapsed = round(time.time() - start_time, 2)
    draft_progress["done"] = True
    draft_progress["running"] = False

    return AgentResult(
        agent_name="DraftingAgent",
        status=AgentStatus.SUCCESS,
        data={
            "total_questions": draft_progress["total"],
            "exact_matched": exact_count,
            "fuzzy_matched": fuzzy_count,
            "total_drafted": drafted_count,
            "warnings": warnings,
        },
        processing_time_sec=elapsed,
    )
