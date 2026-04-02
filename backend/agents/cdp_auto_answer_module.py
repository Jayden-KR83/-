"""
CDP Auto Answer Generation Module
Excel Comm. Tool_SKEP 시트 기반 CDP 1st Answer 자동 생성

Phase 1: 규칙 기반 (BQ Close 스킵, 전년도 복사, 고정 선택형)
Phase 2: AI 초안 생성 (Claude API - get_claude_client() 연동)
"""
import re
import logging
from dataclasses import dataclass, field
from enum import Enum
from typing import Optional, Dict, List, Any

from backend.core.llm_client import get_claude_client
from backend.core.config import settings

logger = logging.getLogger(__name__)


# ── 답변 유형 ─────────────────────────────────────────────────────────────────

class AnswerType(Enum):
    SINGLE_SELECT  = "single_select"
    MULTI_SELECT   = "multi_select"
    TEXT_SHORT     = "text_short"      # ≤ 1500자
    TEXT_LONG      = "text_long"       # > 1500자
    NUMERICAL      = "numerical"
    DATE           = "date"
    PERCENTAGE     = "percentage"
    ATTACHMENT     = "attachment"
    GROUPED_OPTION = "grouped_option"
    SKIP           = "skip"            # BQ = "X" 또는 빈 옵션

class AutoLevel(Enum):
    FULL_AUTO        = "full_auto"       # 신뢰도 ≥ 0.95, 자동 저장 가능
    AI_DRAFT         = "ai_draft"        # AI 생성 초안, 담당자 검수 필요
    MANUAL_REQUIRED  = "manual"          # 자동화 불가, 수작업 필요


# ── 문항 데이터 구조 ──────────────────────────────────────────────────────────

@dataclass
class CDPQuestion:
    row_idx:          int
    order:            str
    q_no:             str
    level:            str
    q_type:           str
    main_question:    str        # U  (English)
    main_question_kr: str        # T  (Korean)
    sub_q_no:         str
    sub_q_desc:       str
    sub_q_desc_kr:    str
    sub_q_options:    str
    sub_q_options_kr: str
    open_close:       str
    row_group:        str
    prev_year_answer: str        # BZ (2024 CDP)
    current_answer:   str        # BX (2025 CDP 1st_Ans., existing)
    final_answer:     str        # CB (2025 CDP_Final_Ans.)
    scoring_dc:       str        # W
    scoring_ac:       str        # Y
    scoring_mc:       str        # AA
    scoring_lc:       str        # AC
    # CDP 공식 채점 (PC group: AF~AM)
    d_den_cdp:        str        # AF D_Den_CDP
    a_den_cdp:        str        # AG A_Den_CDP
    m_den_cdp:        str        # AH M_Den_CDP
    l_den_cdp:        str        # AI L_Den_CDP
    d_num_cdp:        str        # AJ D_Num_CDP
    a_num_cdp:        str        # AK A_Num_CDP
    m_num_cdp:        str        # AL M_Num_CDP
    l_num_cdp:        str        # AM L_Num_CDP
    # 부가 정보
    rationale:        str        # BK
    ambition:         str        # BM
    response_options: str        # BO
    dependencies:     str
    change_status:    str
    remarks:          str        # BY
    recheck:          str        # CA
    cold_eye:         str        # CC
    answer_type:      Optional[AnswerType] = None
    auto_level:       Optional[AutoLevel]  = None
    generated_answer: Optional[str]        = None
    confidence:       float                = 0.0

    def to_dict(self) -> Dict[str, Any]:
        """API 직렬화용 dict 변환"""
        return {
            "row_idx":          self.row_idx,
            "order":            self.order,
            "q_no":             self.q_no,
            "level":            self.level,
            "q_type":           self.q_type,
            "main_question":    self.main_question,
            "main_question_kr": self.main_question_kr,
            "sub_q_no":         self.sub_q_no,
            "sub_q_desc":       self.sub_q_desc,
            "sub_q_desc_kr":    self.sub_q_desc_kr,
            "sub_q_options":    self.sub_q_options,
            "sub_q_options_kr": self.sub_q_options_kr,
            "open_close":       self.open_close,
            "row_group":        self.row_group,
            "prev_year_answer": self.prev_year_answer,
            "current_answer":   self.current_answer,
            "final_answer":     self.final_answer,
            "scoring_dc":       self.scoring_dc,
            "scoring_ac":       self.scoring_ac,
            "scoring_mc":       self.scoring_mc,
            "scoring_lc":       self.scoring_lc,
            "d_den_cdp":        self.d_den_cdp,
            "a_den_cdp":        self.a_den_cdp,
            "m_den_cdp":        self.m_den_cdp,
            "l_den_cdp":        self.l_den_cdp,
            "d_num_cdp":        self.d_num_cdp,
            "a_num_cdp":        self.a_num_cdp,
            "m_num_cdp":        self.m_num_cdp,
            "l_num_cdp":        self.l_num_cdp,
            "rationale":        self.rationale,
            "ambition":         self.ambition,
            "response_options": self.response_options,
            "dependencies":     self.dependencies,
            "change_status":    self.change_status,
            "remarks":          self.remarks,
            "recheck":          self.recheck,
            "cold_eye":         self.cold_eye,
            "answer_type":      self.answer_type.value  if self.answer_type  else None,
            "auto_level":       self.auto_level.value   if self.auto_level   else None,
            "generated_answer": self.generated_answer,
            "confidence":       self.confidence,
            "has_answer":       bool(self.current_answer or self.final_answer),
            "has_prev":         bool(self.prev_year_answer),
            "is_scored":        bool(self.scoring_dc and self.scoring_dc.strip() != "Not scored."),
            "is_open":          self.open_close != "X",
        }


# ── 답변 유형 분류기 ──────────────────────────────────────────────────────────

class AnswerTypeClassifier:
    ATTACHMENT_P  = [r"\[Attachment"]
    DATE_P        = [r"Date field", r"\[DD/MM/YYYY\]"]
    PERCENTAGE_P  = [r"Percentage field", r"enter a percentage"]
    NUMERICAL_P   = [r"Numerical field", r"enter a number from"]
    GROUPED_P     = [r"Grouped option", r"multi-select group"]
    MULTI_P       = [r"Select all that apply"]
    SINGLE_P      = [r"^Select from:"]
    TEXT_P        = [r"Text field \[maximum ([\d,]+) characters?\]"]

    @classmethod
    def classify(cls, q: CDPQuestion) -> AnswerType:
        if q.open_close == "X":
            return AnswerType.SKIP
        opt = (q.sub_q_options or "").strip()
        if not opt:
            return AnswerType.SKIP
        for patterns, atype in [
            (cls.ATTACHMENT_P, AnswerType.ATTACHMENT),
            (cls.DATE_P,       AnswerType.DATE),
            (cls.PERCENTAGE_P, AnswerType.PERCENTAGE),
            (cls.NUMERICAL_P,  AnswerType.NUMERICAL),
            (cls.GROUPED_P,    AnswerType.GROUPED_OPTION),
            (cls.MULTI_P,      AnswerType.MULTI_SELECT),
            (cls.SINGLE_P,     AnswerType.SINGLE_SELECT),
        ]:
            for p in patterns:
                if re.search(p, opt, re.IGNORECASE):
                    return atype
        for p in cls.TEXT_P:
            m = re.search(p, opt, re.IGNORECASE)
            if m:
                mc = int(m.group(1).replace(",", ""))
                return AnswerType.TEXT_SHORT if mc <= 1500 else AnswerType.TEXT_LONG
        return AnswerType.TEXT_SHORT

    @classmethod
    def determine_auto_level(cls, q: CDPQuestion, atype: AnswerType) -> AutoLevel:
        if atype == AnswerType.SKIP:
            return AutoLevel.FULL_AUTO
        if atype == AnswerType.ATTACHMENT:
            return AutoLevel.MANUAL_REQUIRED
        has_prev   = bool(q.prev_year_answer and q.prev_year_answer.strip())
        no_change  = q.change_status in ("No change", "", None)
        # 전년도 답변 있고 변경 없는 선택/수치형 → 완전 자동
        if has_prev and no_change and atype in (
            AnswerType.SINGLE_SELECT, AnswerType.MULTI_SELECT,
            AnswerType.DATE, AnswerType.NUMERICAL, AnswerType.PERCENTAGE,
        ):
            return AutoLevel.FULL_AUTO
        if atype == AnswerType.TEXT_LONG:
            return AutoLevel.AI_DRAFT if has_prev else AutoLevel.MANUAL_REQUIRED
        if atype in (AnswerType.TEXT_SHORT, AnswerType.GROUPED_OPTION):
            return AutoLevel.AI_DRAFT
        return AutoLevel.AI_DRAFT


# ── Phase 1: 규칙 기반 엔진 ───────────────────────────────────────────────────

class RuleBasedEngine:
    """
    비용 0, 즉시 적용 가능.
    Close 스킵 + 전년도 복사 + 회사 고정값으로 약 40~50% 자동 처리.
    """
    COMPANY_PROFILE = {
        "language":                      "English",
        "currency":                      "KRW",
        "organization_type":             "Privately owned organization",
        "country":                       "Republic of Korea",
        "financial_reporting_alignment": "Yes",
        "reporting_boundary_same":       "Yes",
        "unique_identifier_type":        "LEI number",
        "uses_unique_identifier":        "Yes",
        "lei_number":                    "988400DY41E2FT2XK563",
    }

    # q_no → 고정 답변 (변경 빈도 낮은 행정 정보)
    FIXED_ANSWERS: Dict[str, str] = {
        "[1.1]": "English",
        "[1.2]": "KRW",
    }

    @classmethod
    def generate(cls, q: CDPQuestion) -> Optional[str]:
        """
        반환값:
          - str  : 자동 생성 답변
          - ""   : Close / 빈칸 처리
          - None : 규칙으로 처리 불가 → Phase 2로 위임
        """
        if q.open_close == "X":
            return ""
        if q.q_no in cls.FIXED_ANSWERS:
            return cls.FIXED_ANSWERS[q.q_no]
        has_prev  = bool(q.prev_year_answer and q.prev_year_answer.strip())
        no_change = q.change_status in ("No change", "", None)
        if has_prev and no_change and q.answer_type in (
            AnswerType.SINGLE_SELECT, AnswerType.MULTI_SELECT,
            AnswerType.DATE, AnswerType.NUMERICAL, AnswerType.PERCENTAGE,
        ):
            return q.prev_year_answer.strip()
        return None  # Phase 2로


# ── Phase 2: AI 초안 생성 엔진 ────────────────────────────────────────────────

class AIAnswerGenerator:
    """
    기존 Platform의 get_claude_client() 사용.
    Claude API 비용 발생 — enable_ai=False로 Phase 1만 먼저 테스트 권장.
    """

    @classmethod
    def build_prompt(cls, q: CDPQuestion, company_context: str = "") -> str:
        parts = []
        if q.main_question:
            parts.append(f"## Main Question ({q.q_no})\n{q.main_question}")
        if q.sub_q_desc:
            parts.append(f"## Sub Question (#{q.sub_q_no})\nEN: {q.sub_q_desc}\nKR: {q.sub_q_desc_kr}")
        if q.sub_q_options:
            parts.append(f"## Available Options\n{q.sub_q_options}")
        scoring = []
        for label, val in [("D", q.scoring_dc), ("A", q.scoring_ac),
                           ("M", q.scoring_mc), ("L", q.scoring_lc)]:
            if val and val.strip() and val != "Not scored.":
                scoring.append(f"[{label}] {val}")
        if scoring:
            parts.append("## Scoring Criteria\n" + "\n".join(scoring))
        if q.prev_year_answer:
            parts.append(f"## Previous Year Answer\n{q.prev_year_answer[:2000]}")
        if q.change_status and q.change_status != "No change":
            parts.append(f"## Change from Last Year: {q.change_status}")
        if q.dependencies:
            parts.append(f"## Dependencies\n{q.dependencies}")
        if company_context:
            parts.append(f"## Company Context\n{company_context}")
        # 출력 형식 지시
        at = q.answer_type
        if at == AnswerType.SINGLE_SELECT:
            parts.append("## Output: Return EXACTLY one option from the available list above.")
        elif at == AnswerType.MULTI_SELECT:
            parts.append("## Output: Return all applicable options, each prefixed with bullet (•).")
        elif at in (AnswerType.TEXT_SHORT, AnswerType.TEXT_LONG):
            mc = cls._extract_max_chars(q.sub_q_options)
            parts.append(f"## Output: Write comprehensive answer within {mc} characters.")
        elif at == AnswerType.NUMERICAL:
            parts.append("## Output: Return ONLY a number (no units, no commas).")
        elif at == AnswerType.PERCENTAGE:
            parts.append("## Output: Return percentage value only (0-100, up to 2 decimals).")
        elif at == AnswerType.DATE:
            parts.append("## Output: Return date in DD/MM/YYYY format only.")
        elif at == AnswerType.GROUPED_OPTION:
            parts.append("## Output: Return groups with selected options using ☑ prefix.")
        return "\n\n".join(parts)

    @staticmethod
    def _extract_max_chars(options_text: str) -> int:
        m = re.search(r"maximum ([\d,]+) characters?", options_text or "", re.IGNORECASE)
        return int(m.group(1).replace(",", "")) if m else 2500

    @classmethod
    def call_claude_api(cls, prompt: str, max_tokens: int = 4096) -> str:
        """기존 Platform ClaudeClient 사용 — anthropic SDK 직접 호출 없음"""
        client = get_claude_client()
        system = (
            "You are an expert CDP response consultant for SK ecoplant Co., Ltd.\n"
            "Company: SK ecoplant | Industry: Construction/Environmental/Energy | HQ: Republic of Korea\n"
            "Net Zero: 2040 | SBTi: Approved (1.5C aligned) | Revenue: ~9.3T KRW (2024)\n\n"
            "Rules:\n"
            "1) MAXIMIZE CDP score\n"
            "2) For 'Select from:' → pick the single BEST scoring option\n"
            "3) For 'Select all that apply' → pick ALL truthful applicable options\n"
            "4) For Text → include specific data, metrics, named initiatives, TCFD/SBTi/GHG Protocol references\n"
            "5) NEVER fabricate data — flag uncertain items with [VERIFY_REQUIRED]\n"
            "6) Stay within character limits\n"
            "7) Respond in professional English"
        )
        return client.chat(
            prompt,
            system_prompt=system,
            temperature=0.1,
            max_tokens=max_tokens,
        )


# ── 문항 간 의존성(Cross-Reference) 엔진 ─────────────────────────────────────

class DependencyResolver:
    _PATTERNS = [
        re.compile(
            r'select\s+"([^"]+)"\s+in\s+response\s+to\s+(?:.*?)?(\d+\.\d+(?:\.\d+)?)',
            re.IGNORECASE,
        ),
        re.compile(r'only\s+appears\s+if\s+you\s+select\s+"([^"]+)"', re.IGNORECASE),
    ]

    @classmethod
    def parse(cls, dep_text: str) -> List[Dict]:
        if not dep_text:
            return []
        conditions = []
        for pat in cls._PATTERNS:
            for m in pat.findall(dep_text):
                if len(m) >= 2:
                    conditions.append({
                        "required_answer":  m[0],
                        "source_question":  f"[{m[1]}]",
                    })
        return conditions

    @classmethod
    def is_met(cls, q: CDPQuestion, answered: Dict[str, str]) -> bool:
        for cond in cls.parse(q.dependencies):
            src = cond["source_question"]
            if src in answered:
                if cond["required_answer"].lower() not in answered[src].lower():
                    return False
        return True

    @classmethod
    def sorted_order(cls, questions: List[CDPQuestion]) -> List[int]:
        """의존성 없는 문항 먼저, 의존성 있는 문항 뒤로"""
        no_dep  = [q.row_idx for q in questions if not q.dependencies]
        has_dep = [q.row_idx for q in sorted(questions, key=lambda x: x.q_no)
                   if q.dependencies]
        return no_dep + has_dep


# ── 오케스트레이터 (Phase 1 + 2 통합 실행) ───────────────────────────────────

class CDPAutoAnswerOrchestrator:
    def __init__(self, enable_ai: bool = True):
        self.enable_ai  = enable_ai
        self.classifier = AnswerTypeClassifier()
        self.rule       = RuleBasedEngine()
        self.ai         = AIAnswerGenerator()
        self.dep        = DependencyResolver()

    def process_single(
        self,
        q: CDPQuestion,
        company_context: str = "",
        answered: Dict[str, str] = None,
    ) -> CDPQuestion:
        answered = answered or {}
        q.answer_type = self.classifier.classify(q)
        q.auto_level  = self.classifier.determine_auto_level(q, q.answer_type)

        # 의존성 미충족 → 빈칸
        if not self.dep.is_met(q, answered):
            q.generated_answer = ""
            q.confidence       = 1.0
            q.auto_level       = AutoLevel.FULL_AUTO
            return q

        # Phase 1: 규칙 기반
        rule_answer = self.rule.generate(q)
        if rule_answer is not None:
            q.generated_answer = rule_answer
            q.confidence       = 0.95 if rule_answer else 1.0
            return q

        # Phase 2: AI
        if self.enable_ai and q.auto_level != AutoLevel.MANUAL_REQUIRED:
            prompt = self.ai.build_prompt(q, company_context)
            try:
                q.generated_answer = self.ai.call_claude_api(prompt)
                q.confidence = (
                    0.70 if q.answer_type in (AnswerType.TEXT_LONG, AnswerType.GROUPED_OPTION)
                    else 0.85
                )
            except Exception as e:
                logger.warning(f"AI 생성 실패 {q.q_no}#{q.sub_q_no}: {e}")
                q.generated_answer = None
                q.confidence       = 0.0
        else:
            q.generated_answer = None
            q.confidence       = 0.0
        return q

    def process_all(
        self,
        questions: List[CDPQuestion],
        company_context: str = "",
        progress_callback=None,
    ) -> Dict[str, Any]:
        idx_map  = {q.row_idx: q for q in questions}
        order    = self.dep.sorted_order(questions)
        answered: Dict[str, str] = {}
        stats    = {"total": len(questions), "full_auto": 0, "ai_draft": 0,
                    "manual": 0, "skipped": 0}
        out_questions = []

        for i, row_idx in enumerate(order):
            q = idx_map[row_idx]
            q = self.process_single(q, company_context, answered)

            if q.generated_answer:
                answered[q.q_no] = q.generated_answer

            if q.auto_level == AutoLevel.FULL_AUTO:
                key = "skipped" if q.answer_type == AnswerType.SKIP else "full_auto"
                stats[key] += 1
            elif q.auto_level == AutoLevel.AI_DRAFT:
                stats["ai_draft"] += 1
            else:
                stats["manual"] += 1

            out_questions.append({
                "row":          q.row_idx,
                "q_no":         q.q_no,
                "sub_q_no":     q.sub_q_no,
                "sub_q_desc":   q.sub_q_desc_kr or q.sub_q_desc,
                "answer_type":  q.answer_type.value  if q.answer_type  else None,
                "auto_level":   q.auto_level.value   if q.auto_level   else None,
                "answer":       q.generated_answer,
                "confidence":   q.confidence,
                "needs_review": q.confidence < 0.9 and bool(q.generated_answer),
                "open_close":   q.open_close,
            })

            if progress_callback:
                progress_callback(i + 1, len(order))

        return {**stats, "questions": out_questions}


# ── Excel 데이터 로더 ─────────────────────────────────────────────────────────

class ExcelDataLoader:
    """
    Comm. Tool_SKEP 시트 → CDPQuestion 리스트 변환.

    COL_MAP은 Excel 실제 열 순서(0-based index)와 반드시 일치해야 함.
    Excel 구조 변경 시 이 매핑만 수정하면 됨.
    """
    COL_MAP = {
        # ── 문항 기본 정보 ────────────────────────────────────────────
        "order":            0,   # A  순번
        "q_no":             1,   # B  Q_No.
        "level":            2,   # C  Lv.
        "q_type":           17,  # R  질문 유형
        "main_question_kr": 19,  # T  K_Q_Des. (한국어 메인 질문)
        "main_question":    20,  # U  Q_Des.   (영어 메인 질문)
        # ── 채점 기준 (W~AD) ──────────────────────────────────────────
        "scoring_dc":       22,  # W  DC (Disclosure Criteria EN)
        "scoring_ac":       24,  # Y  AC (Awareness Criteria EN)
        "scoring_mc":       26,  # AA MC (Management Criteria EN)
        "scoring_lc":       28,  # AC LC (Leadership Criteria EN)
        # ── CDP 공식 채점 분모/분자 (AF~AM) ──────────────────────────
        "d_den_cdp":        31,  # AF D_Den_CDP
        "a_den_cdp":        32,  # AG A_Den_CDP
        "m_den_cdp":        33,  # AH M_Den_CDP
        "l_den_cdp":        34,  # AI L_Den_CDP
        "d_num_cdp":        35,  # AJ D_Num_CDP
        "a_num_cdp":        36,  # AK A_Num_CDP
        "m_num_cdp":        37,  # AL M_Num_CDP
        "l_num_cdp":        38,  # AM L_Num_CDP
        # ── 부가 정보 (BG~BO) ─────────────────────────────────────────
        "dependencies":     58,  # BG Question_dependencies
        "change_status":    60,  # BI Change_from_last_year
        "rationale":        62,  # BK Rationale
        "ambition":         64,  # BM Ambition
        "response_options": 66,  # BO Response_options
        # ── 서브 질문 / 답변 영역 (BQ~CC) ────────────────────────────
        "open_close":       68,  # BQ Open or Close (O/X)
        "row_group":        69,  # BR Row
        "sub_q_no":         70,  # BS 1_Sub_Q_No.
        "sub_q_desc":       71,  # BT 2_Sub_Q_Des. (영어)
        "sub_q_desc_kr":    72,  # BU K_2_Sub_Q_Des. (한국어)
        "sub_q_options":    73,  # BV 3_Sub_Q_Op (영어 선택지)
        "sub_q_options_kr": 74,  # BW K_3_Sub_Q_Op (한국어)
        "current_answer":   75,  # BX 2025 CDP 1st_Ans. ← 자동화 대상
        "remarks":          76,  # BY Remarks
        "prev_answer":      77,  # BZ 2024 CDP (전년도)
        "recheck":          78,  # CA Re-Check or Not
        "final_answer":     79,  # CB 2025 CDP_Final_Ans.
        "cold_eye":         80,  # CC Remarks(Cold-Eye)
    }

    @classmethod
    def _ensure_xlsx(cls, excel_path: str) -> None:
        """OLE2(.xls) 포맷 파일을 실제 OOXML(.xlsx)로 제자리 변환.
        이미 정상 xlsx이면 아무것도 하지 않는다."""
        import zipfile
        if zipfile.is_zipfile(excel_path):
            return
        try:
            import xlrd, openpyxl
        except ImportError:
            raise ImportError("xlrd/openpyxl 필요: pip install xlrd openpyxl")
        xls_wb  = xlrd.open_workbook(excel_path)
        xlsx_wb = openpyxl.Workbook()
        xlsx_wb.remove(xlsx_wb.active)
        for sname in xls_wb.sheet_names():
            xls_ws  = xls_wb.sheet_by_name(sname)
            xlsx_ws = xlsx_wb.create_sheet(title=sname)
            for r in range(xls_ws.nrows):
                xlsx_ws.append(xls_ws.row_values(r))
        xlsx_wb.save(excel_path)
        logger.info(f"XLS→XLSX 변환 완료: {excel_path}")

    @classmethod
    def load(cls, excel_path: str, sheet_name: str = "Comm. Tool_SKEP") -> List[CDPQuestion]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl 필요: pip install openpyxl")

        cls._ensure_xlsx(excel_path)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트 '{sheet_name}' 없음. 사용 가능: {wb.sheetnames}")
        ws = wb[sheet_name]

        questions = []
        for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
            if not row:
                continue
            q_no_val = row[cls.COL_MAP["q_no"]] if len(row) > cls.COL_MAP["q_no"] else None
            if not q_no_val:
                continue

            def _s(col_key: str) -> str:
                idx = cls.COL_MAP.get(col_key, -1)
                if idx < 0 or idx >= len(row):
                    return ""
                v = row[idx]
                return str(v).strip() if v is not None else ""

            questions.append(CDPQuestion(
                row_idx          = 4 + i,
                order            = _s("order"),
                q_no             = _s("q_no"),
                level            = _s("level"),
                q_type           = _s("q_type"),
                main_question    = _s("main_question"),
                main_question_kr = _s("main_question_kr"),
                sub_q_no         = _s("sub_q_no"),
                sub_q_desc       = _s("sub_q_desc"),
                sub_q_desc_kr    = _s("sub_q_desc_kr"),
                sub_q_options    = _s("sub_q_options"),
                sub_q_options_kr = _s("sub_q_options_kr"),
                open_close       = _s("open_close"),
                row_group        = _s("row_group"),
                prev_year_answer = _s("prev_answer"),
                current_answer   = _s("current_answer"),
                final_answer     = _s("final_answer"),
                scoring_dc       = _s("scoring_dc"),
                scoring_ac       = _s("scoring_ac"),
                scoring_mc       = _s("scoring_mc"),
                scoring_lc       = _s("scoring_lc"),
                d_den_cdp        = _s("d_den_cdp"),
                a_den_cdp        = _s("a_den_cdp"),
                m_den_cdp        = _s("m_den_cdp"),
                l_den_cdp        = _s("l_den_cdp"),
                d_num_cdp        = _s("d_num_cdp"),
                a_num_cdp        = _s("a_num_cdp"),
                m_num_cdp        = _s("m_num_cdp"),
                l_num_cdp        = _s("l_num_cdp"),
                rationale        = _s("rationale"),
                ambition         = _s("ambition"),
                response_options = _s("response_options"),
                dependencies     = _s("dependencies"),
                change_status    = _s("change_status"),
                remarks          = _s("remarks"),
                recheck          = _s("recheck"),
                cold_eye         = _s("cold_eye"),
            ))
        logger.info(f"ExcelDataLoader: {len(questions)}개 문항 로드 ({excel_path})")
        return questions

    @classmethod
    def verify_columns(cls, excel_path: str, sheet_name: str = "Comm. Tool_SKEP") -> Dict[str, Any]:
        """
        COL_MAP이 실제 Excel 열 헤더와 일치하는지 검증.
        Excel 파일 교체 시 먼저 이 함수 실행 권장.
        """
        import openpyxl
        cls._ensure_xlsx(excel_path)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        mismatches = []
        for key, idx in cls.COL_MAP.items():
            if idx < len(headers):
                mismatches.append({
                    "field": key,
                    "col_index": idx,
                    "col_letter": chr(65 + idx) if idx < 26 else f"A{chr(65 + idx - 26)}",
                    "header_value": headers[idx],
                })
        return {"total_columns": len(headers), "mapped_fields": mismatches}


# ── 결과 리포터 ───────────────────────────────────────────────────────────────

class ResultReporter:
    @staticmethod
    def summary(results: Dict) -> str:
        t  = results["total"]
        fa = results["full_auto"]
        ai = results["ai_draft"]
        m  = results["manual"]
        sk = results["skipped"]
        return (
            f"\n{'='*44}\n"
            f" CDP Auto Answer Report\n"
            f"{'='*44}\n"
            f" 전체: {t}  |  완전자동: {fa}  |  AI초안: {ai}\n"
            f" 수작업: {m}  |  건너뜀: {sk}\n"
            f" 규칙기반 처리율: {(fa+sk)/t*100:.1f}%\n"
            f" AI포함 처리율:   {(fa+ai+sk)/t*100:.1f}%\n"
            f"{'='*44}"
        )

    @staticmethod
    def review_list(results: Dict) -> List[Dict]:
        return [
            {
                "row":         q["row"],
                "q_no":        q["q_no"],
                "sub_q_no":    q["sub_q_no"],
                "answer_type": q["answer_type"],
                "confidence":  q["confidence"],
                "draft":       (q["answer"] or "")[:200],
            }
            for q in results["questions"]
            if q["needs_review"] and q["answer"]
        ]


# ── 플랫폼 통합 엔트리포인트 ─────────────────────────────────────────────────

def run_cdp_auto_answer(
    excel_path: str,
    sheet_name: str        = "Comm. Tool_SKEP",
    company_context: str   = "",
    enable_ai: bool        = False,   # 기본 False: Phase 1만 (비용 0)
    output_mode: str       = "preview",  # "preview" | "write"
    target_questions: List[str] = None,   # None=전체, 지정 시 해당 q_no만
    progress_callback      = None,
) -> Dict[str, Any]:
    """
    CDP 1st Answer 자동 생성 메인 함수.

    Args:
        excel_path:        CDP Master Excel 파일 경로
        sheet_name:        대상 시트 (기본: "Comm. Tool_SKEP")
        company_context:   회사 추가 컨텍스트 (AI 생성 시 사용)
        enable_ai:         True = Phase 1 + 2, False = Phase 1만
        output_mode:       "preview" = 결과 반환만, "write" = BX열에 직접 저장
        target_questions:  특정 q_no 리스트 (예: ["7.1", "7.2"])
        progress_callback: fn(current, total) 진행률 콜백

    Returns:
        results dict: total/full_auto/ai_draft/manual/skipped/questions
    """
    # 1. Excel 로드
    questions = ExcelDataLoader.load(excel_path, sheet_name)

    # 2. 대상 문항 필터
    if target_questions:
        q_set     = {f"[{q}]" for q in target_questions} | set(target_questions)
        questions = [q for q in questions if q.q_no in q_set]
        logger.info(f"대상 문항 필터: {len(questions)}개")

    # 3. 자동 생성 실행
    orchestrator = CDPAutoAnswerOrchestrator(enable_ai=enable_ai)
    results      = orchestrator.process_all(questions, company_context, progress_callback)

    logger.info(ResultReporter.summary(results))

    # 4. 결과 저장 (write 모드)
    if output_mode == "write":
        import openpyxl
        ExcelDataLoader._ensure_xlsx(excel_path)
        wb2  = openpyxl.load_workbook(excel_path)
        ws2  = wb2[sheet_name]
        col  = ExcelDataLoader.COL_MAP["current_answer"] + 1  # 1-based
        written = 0
        for q in results["questions"]:
            if q["answer"] and q["confidence"] >= 0.9:
                ws2.cell(row=q["row"], column=col).value = q["answer"]
                written += 1
        wb2.save(excel_path)
        logger.info(f"Excel 저장 완료: {written}개 답변 BX열에 기록")
        results["written"] = written

    return results
