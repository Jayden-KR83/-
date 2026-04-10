# -*- coding: utf-8 -*-
"""
Communication Tool — CDP Master Analysis Sheet Builder
=======================================================
Merges Source A (Questionnaire parsed) and Source B (Scoring parsed)
into a single CDP_2025_Master_Analysis.xlsx with columns A–AF.

Column Layout (A–AF, 32 columns):
  A  Order_No        — unique sequential number per main question
  B  Q_No            — question ID (e.g. 1.3, 2.1a)
  C  Lv              — "Lv.1" (main) / "Lv.2" (sub)
  D  Q_Des           — English question description (Lv.1 only)
  E  K_Q_Des         — Korean translation of D
  F  DC              — Disclosure criteria (English)
  G  K_DC            — Korean translation of F
  H  AC              — Awareness criteria (English)
  I  K_AC            — Korean translation of H
  J  MC              — Management criteria (English)
  K  K_MC            — Korean translation of J
  L  LC              — Leadership criteria (English)
  M  K_LC            — Korean translation of L
  N  Max_Points      — 최대배점
  O  Sector          — 섹터 정보
  P  D_Num           — Disclosure numerator
  Q  D_Den           — Disclosure denominator
  R  A_Num           — Awareness numerator
  S  A_Den           — Awareness denominator
  T  M_Num           — Management numerator
  U  M_Den           — Management denominator
  V  L_Num           — Leadership numerator
  W  L_Den           — Leadership denominator
  X  Open_or_Close   — user input (O / X / O or X)
  Y  Sub_No          — sub-question number (Lv.2 only)
  Z  Sub_Des         — sub-question description (Lv.2 only)
  AA Sub_Options     — options text (Lv.2 only)
  AB K_Sub_Des       — Korean translation of Z
  AC K_Sub_Options   — Korean translation of AA
  AD Answer          — placeholder
  AE Review          — placeholder
  AF Note            — placeholder
"""

import re
import time
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation

from backend.core.models import AgentResult, AgentStatus

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MASTER_COLUMNS = [
    "Order_No", "Q_No", "Lv", "Q_Des", "K_Q_Des",
    "DC", "K_DC", "AC", "K_AC", "MC", "K_MC", "LC", "K_LC",
    "Max_Points", "Sector",
    "D_Num", "D_Den", "A_Num", "A_Den",
    "M_Num", "M_Den", "L_Num", "L_Den",
    "Open_or_Close",
    "Sub_No", "Sub_Des", "Sub_Options",
    "K_Sub_Des", "K_Sub_Options",
    "Answer", "Review", "Note",
]

SOURCE_A_FILES = [
    "data/outputs/GR A_Full_Corporate_Questionnaire_Modules_1-6_parsed.xlsx",
    "data/outputs/GR A_Full_Corporate_Questionnaire_Module_7_parsed.xlsx",
    "data/outputs/GR A_Full_Corporate_Questionnaire_Modules_8-13_parsed.xlsx",
]

SOURCE_B_FILE = (
    "data/outputs/"
    "GR A_CDP_Full_Corporate_Scoring_Methodology_2025_-_"
    "Climate_change_v1.0(20250430)_scoring.xlsx"
)

OUTPUT_PATH = "data/outputs/CDP_2025_Master_Analysis.xlsx"

# ---------------------------------------------------------------------------
# Question ID parsing / sorting
# ---------------------------------------------------------------------------
_QID_RE = re.compile(r"(\d+)\.(\d+)(?:\.(\d+))?([a-z])?")


def _parse_qid(qid: str) -> Tuple:
    """Return a sortable tuple from a question ID like '1.3', '1.4.1', '2.1a'."""
    m = _QID_RE.match(qid.strip())
    if not m:
        return (999, 999, 999, "z")
    module = int(m.group(1))
    main = int(m.group(2))
    sub = int(m.group(3)) if m.group(3) else 0
    letter = m.group(4) or ""
    return (module, main, sub, letter)


def _normalize_qid(raw: str) -> str:
    """Strip parentheses and whitespace: '(1.3)' -> '1.3'."""
    return raw.strip().strip("()")


# ---------------------------------------------------------------------------
# Load helpers
# ---------------------------------------------------------------------------
def _load_source_a(base_dir: str = ".") -> pd.DataFrame:
    """Load and combine all 3 questionnaire parsed files."""
    frames = []
    # Try configured paths first, then scan outputs dir
    search_paths = [Path(base_dir) / rel for rel in SOURCE_A_FILES]
    # Also scan data/outputs/ for *Questionnaire*_parsed.xlsx
    outputs_dir = Path(base_dir) / "data" / "outputs"
    if outputs_dir.exists():
        for f in outputs_dir.glob("*Questionnaire*_parsed.xlsx"):
            if f not in search_paths:
                search_paths.append(f)

    for path in search_paths:
        if not path.exists():
            continue
        df = pd.read_excel(str(path), engine="openpyxl", dtype=str)
        df = df.fillna("")
        frames.append(df)
        logger.info("Source A loaded: %s (%d rows)", path.name, len(df))

    if not frames:
        raise FileNotFoundError("Source A 파일을 찾을 수 없습니다. data/outputs/ 에 Questionnaire 파싱 결과가 필요합니다.")
    return pd.concat(frames, ignore_index=True)


def _load_source_b(base_dir: str = ".") -> pd.DataFrame:
    """Load the scoring parsed file."""
    path = Path(base_dir) / SOURCE_B_FILE
    # Fallback: scan outputs for *_scoring.xlsx
    if not path.exists():
        outputs_dir = Path(base_dir) / "data" / "outputs"
        if outputs_dir.exists():
            scoring_files = list(outputs_dir.glob("*_scoring.xlsx"))
            if scoring_files:
                path = scoring_files[0]
    if not path.exists():
        raise FileNotFoundError("Source B 파일을 찾을 수 없습니다. data/outputs/ 에 Scoring 파싱 결과가 필요합니다.")
    df = pd.read_excel(str(path), dtype=str)
    df = df.fillna("")
    logger.info("Source B loaded: %s (%d rows)", path.name, len(df))
    return df


# ---------------------------------------------------------------------------
# Build question groups from Source A
# ---------------------------------------------------------------------------
def _build_question_groups(sa: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    """
    Group Source A rows by question ID.

    For duplicate question IDs (TOC entry vs detail entry), keep the one
    that has sub-rows (번호행) following it.

    Returns dict: qid -> {
        "질문내용": str,
        "subs": [{"번호": ..., "Sub질문": ..., "Options": ...}, ...],
        "tags": [{"Tags_항목": ..., "Tags_내용": ...}, ...],
    }
    """
    groups: Dict[str, Dict[str, Any]] = {}
    current_qid = None
    current_data: Optional[Dict] = None

    for _, row in sa.iterrows():
        if row["행_유형"] == "질문":
            # Save previous
            if current_qid and current_data:
                existing = groups.get(current_qid)
                # Keep the version with more sub-rows
                if existing is None or len(current_data["subs"]) >= len(existing["subs"]):
                    groups[current_qid] = current_data

            qid = _normalize_qid(str(row["질문번호"]))
            current_qid = qid
            current_data = {
                "질문내용": _clean(row.get("질문내용", "")),
                "subs": [],
                "tags": [],
            }
        elif row["행_유형"] == "번호행" and current_data is not None:
            current_data["subs"].append({
                "번호": row.get("번호", ""),
                "Sub질문": _clean(row.get("Sub질문", "")),
                "Options": _clean(row.get("Options", "")),
            })
        elif row["행_유형"] == "태그행" and current_data is not None:
            current_data["tags"].append({
                "Tags_항목": _clean(row.get("Tags_항목", "")),
                "Tags_내용": _clean(row.get("Tags_내용", "")),
            })

    # Save last group
    if current_qid and current_data:
        existing = groups.get(current_qid)
        if existing is None or len(current_data["subs"]) >= len(existing["subs"]):
            groups[current_qid] = current_data

    return groups


def _clean(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("nan", "none") else s


# ---------------------------------------------------------------------------
# Build scoring lookup from Source B
# ---------------------------------------------------------------------------
def _build_scoring_lookup(sb: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    """
    Build a dict: qid -> scoring data.
    For duplicate qids in scoring, merge by appending criteria.
    """
    lookup: Dict[str, Dict[str, Any]] = {}
    for _, row in sb.iterrows():
        qid = _clean(row.get("문항ID", ""))
        if not qid:
            continue
        entry = {
            "DC": _clean(row.get("Disclosure_기준", "")),
            "DC_배점": _clean(row.get("Disclosure_배점", "")),
            "AC": _clean(row.get("Awareness_기준", "")),
            "AC_배점": _clean(row.get("Awareness_배점", "")),
            "MC": _clean(row.get("Management_기준", "")),
            "MC_배점": _clean(row.get("Management_배점", "")),
            "LC": _clean(row.get("Leadership_기준", "")),
            "LC_배점": _clean(row.get("Leadership_배점", "")),
            "Max_Points": _clean(row.get("최대배점", "")),
            "Sector": _clean(row.get("섹터", "")),
            "D_num": _clean(row.get("D_num", "")),
            "D_den": _clean(row.get("D_den", "")),
            "A_num": _clean(row.get("A_num", "")),
            "A_den": _clean(row.get("A_den", "")),
            "M_num": _clean(row.get("M_num", "")),
            "M_den": _clean(row.get("M_den", "")),
            "L_num": _clean(row.get("L_num", "")),
            "L_den": _clean(row.get("L_den", "")),
        }
        if qid not in lookup:
            lookup[qid] = entry
        else:
            # Merge: append criteria text if different
            existing = lookup[qid]
            for key in ("DC", "AC", "MC", "LC"):
                if entry[key] and entry[key] != existing[key]:
                    existing[key] = (existing[key] + "\n" + entry[key]).strip()
    return lookup


# ---------------------------------------------------------------------------
# Korean translation via Anthropic API
# ---------------------------------------------------------------------------
def _translate_batch(texts: List[str], batch_size: int = 20) -> List[str]:
    """
    Translate English texts to Korean using ClaudeClient.
    Falls back to empty string on failure.
    """
    if not texts:
        return []

    try:
        from backend.core.llm_client import ClaudeClient
        client = ClaudeClient()
    except Exception as e:
        logger.warning("ClaudeClient init failed, skipping translation: %s", e)
        return [""] * len(texts)

    results = [""] * len(texts)
    for i in range(0, len(texts), batch_size):
        batch = texts[i:i + batch_size]
        numbered = "\n".join(f"[{j+1}] {t}" for j, t in enumerate(batch))
        prompt = (
            "Translate the following CDP questionnaire texts from English to Korean.\n"
            "Return ONLY the translations, one per line, numbered like [1] ... [2] ...\n"
            "Keep technical CDP terms as-is (e.g. Scope 1, GHG, TCFD).\n\n"
            + numbered
        )
        try:
            response = client.chat(
                prompt=prompt,
                system_prompt="You are a professional CDP/ESG translator. Translate accurately and concisely.",
                temperature=0.1,
                max_tokens=4000,
            )
            # Parse response
            for line in response.strip().splitlines():
                m = re.match(r"\[(\d+)\]\s*(.*)", line.strip())
                if m:
                    idx = int(m.group(1)) - 1 + i
                    if 0 <= idx < len(results):
                        results[idx] = m.group(2).strip()
        except Exception as e:
            logger.warning("Translation batch failed: %s", e)
    return results


# ---------------------------------------------------------------------------
# Main merge logic
# ---------------------------------------------------------------------------
def build_master_sheet(
    base_dir: str = ".",
    output_path: Optional[str] = None,
    translate: bool = True,
) -> AgentResult:
    """
    Build the CDP_2025_Master_Analysis.xlsx by merging Source A and Source B.

    Args:
        base_dir: project root directory
        output_path: override output file path
        translate: whether to translate empty Korean fields via API
    """
    start_time = time.time()
    warnings_list: List[str] = []

    out_path = Path(base_dir) / (output_path or OUTPUT_PATH)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # --- Load data ---
    try:
        sa = _load_source_a(base_dir)
        sb = _load_source_b(base_dir)
    except FileNotFoundError as e:
        return AgentResult(
            agent_name="CommunicationTool",
            status=AgentStatus.FAILED,
            error_message=str(e),
            data={},
            processing_time_sec=round(time.time() - start_time, 2),
        )

    logger.info("Source A loaded: %d rows", len(sa))
    logger.info("Source B loaded: %d rows", len(sb))

    # --- Build lookups ---
    q_groups = _build_question_groups(sa)
    scoring = _build_scoring_lookup(sb)

    logger.info("Question groups: %d, Scoring entries: %d", len(q_groups), len(scoring))

    # --- Collect all unique question IDs from both sources ---
    all_qids = sorted(
        set(q_groups.keys()) | set(scoring.keys()),
        key=_parse_qid,
    )

    # --- All Q_IDs from both sources are Lv.1 main questions ---
    # Lv.2 sub-rows come from Source A's "번호행" (sub-questions within a question)
    # 3-level IDs like 1.4.1 are still independent main questions, NOT Lv.2

    # --- Build master rows (stable sort: main question, then its subs) ---
    master_rows: List[Dict[str, Any]] = []
    order_no = 0

    for qid in all_qids:
        q_data = q_groups.get(qid, {})
        s_data = scoring.get(qid, {})

        order_no += 1

        # --- Lv.1 row (main question — every Q_ID is a main question) ---
        lv1_row = {col: "" for col in MASTER_COLUMNS}
        lv1_row["Order_No"] = str(order_no)
        lv1_row["Q_No"] = qid
        lv1_row["Lv"] = "Lv.1"
        lv1_row["Q_Des"] = q_data.get("질문내용", s_data.get("질문내용", ""))

        # Scoring criteria
        lv1_row["DC"] = s_data.get("DC", "")
        lv1_row["AC"] = s_data.get("AC", "")
        lv1_row["MC"] = s_data.get("MC", "")
        lv1_row["LC"] = s_data.get("LC", "")
        lv1_row["Max_Points"] = s_data.get("Max_Points", "")
        lv1_row["Sector"] = s_data.get("Sector", "")

        # Point allocation
        lv1_row["D_Num"] = s_data.get("D_num", "")
        lv1_row["D_Den"] = s_data.get("D_den", "")
        lv1_row["A_Num"] = s_data.get("A_num", "")
        lv1_row["A_Den"] = s_data.get("A_den", "")
        lv1_row["M_Num"] = s_data.get("M_num", "")
        lv1_row["M_Den"] = s_data.get("M_den", "")
        lv1_row["L_Num"] = s_data.get("L_num", "")
        lv1_row["L_Den"] = s_data.get("L_den", "")

        # If this question has no subs from Source A, just add the Lv.1 row
        subs = q_data.get("subs", [])
        if not subs:
            master_rows.append(lv1_row)
        else:
            # Add Lv.1 row first
            master_rows.append(lv1_row)

            # Then add Lv.2 rows for each sub-question
            for sub in subs:
                lv2_row = {col: "" for col in MASTER_COLUMNS}
                lv2_row["Order_No"] = str(order_no)
                lv2_row["Q_No"] = qid
                lv2_row["Lv"] = "Lv.2"
                lv2_row["Sub_No"] = str(_clean(sub.get("번호", "")))
                lv2_row["Sub_Des"] = _clean(sub.get("Sub질문", ""))
                lv2_row["Sub_Options"] = _clean(sub.get("Options", ""))
                master_rows.append(lv2_row)

    df = pd.DataFrame(master_rows, columns=MASTER_COLUMNS)
    logger.info("Master sheet built: %d rows x %d cols", len(df), len(df.columns))

    # --- Korean translation ---
    # First, use existing Korean data from Source B if available
    # Source B already has Korean columns? No — criteria are in English.
    # We translate: Q_Des -> K_Q_Des, DC -> K_DC, etc.
    if translate:
        _apply_translations(df, warnings_list)

    # --- Save to Excel ---
    try:
        _save_master_excel(df, str(out_path))
    except Exception as e:
        return AgentResult(
            agent_name="CommunicationTool",
            status=AgentStatus.FAILED,
            error_message="Excel 저장 실패: %s" % str(e),
            data={},
            processing_time_sec=round(time.time() - start_time, 2),
        )

    elapsed = round(time.time() - start_time, 2)
    return AgentResult(
        agent_name="CommunicationTool",
        status=AgentStatus.SUCCESS,
        data={
            "output_path": str(out_path),
            "total_rows": len(df),
            "main_questions": len(df[df["Lv"] == "Lv.1"]),
            "sub_rows": len(df[df["Lv"] == "Lv.2"]),
            "warnings": warnings_list,
        },
        processing_time_sec=elapsed,
    )


# ---------------------------------------------------------------------------
# Translation application
# ---------------------------------------------------------------------------
def _apply_translations(df: pd.DataFrame, warnings: List[str]) -> None:
    """Translate empty Korean columns using Anthropic API."""
    pairs = [
        ("Q_Des", "K_Q_Des"),
        ("DC", "K_DC"),
        ("AC", "K_AC"),
        ("MC", "K_MC"),
        ("LC", "K_LC"),
        ("Sub_Des", "K_Sub_Des"),
        ("Sub_Options", "K_Sub_Options"),
    ]

    for src_col, dst_col in pairs:
        # Find rows where source is non-empty but destination is empty
        mask = (df[src_col].astype(str).str.strip() != "") & (
            df[dst_col].astype(str).str.strip().isin(["", "nan"])
        )
        indices = df.index[mask].tolist()
        if not indices:
            continue

        texts = df.loc[indices, src_col].astype(str).tolist()

        # Deduplicate to save API calls
        unique_texts = list(dict.fromkeys(texts))
        logger.info(
            "Translating %d unique texts for %s -> %s",
            len(unique_texts), src_col, dst_col,
        )

        if len(unique_texts) > 500:
            warnings.append(
                f"Too many texts to translate for {dst_col} ({len(unique_texts)}), "
                "truncating to 500. Remaining will be empty."
            )
            unique_texts = unique_texts[:500]

        translations = _translate_batch(unique_texts)
        trans_map = dict(zip(unique_texts, translations))

        for idx in indices:
            src_text = str(df.at[idx, src_col])
            translated = trans_map.get(src_text, "")
            if translated:
                df.at[idx, dst_col] = translated


# ---------------------------------------------------------------------------
# Excel save with formatting and validation
# ---------------------------------------------------------------------------
def _save_master_excel(df: pd.DataFrame, path: str) -> None:
    """Save DataFrame to Excel with styling and data validation."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="CDP_Master", index=False)

        wb = writer.book
        ws = wb["CDP_Master"]

        # --- Header styling ---
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="center")
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = wrap_align
            cell.border = thin_border

        # --- Column widths ---
        col_widths = {
            "A": 8, "B": 10, "C": 6, "D": 50, "E": 50,
            "F": 40, "G": 40, "H": 40, "I": 40, "J": 40, "K": 40,
            "L": 40, "M": 40, "N": 10, "O": 25,
            "P": 8, "Q": 8, "R": 8, "S": 8, "T": 8, "U": 8, "V": 8, "W": 8,
            "X": 12,
            "Y": 8, "Z": 40, "AA": 50, "AB": 40, "AC": 50,
            "AD": 20, "AE": 20, "AF": 20,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # --- Row styling ---
        lv1_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        lv2_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        body_align = Alignment(wrap_text=True, vertical="top")

        for row_idx in range(2, len(df) + 2):
            lv_cell = ws.cell(row=row_idx, column=3)  # Column C = Lv
            fill = lv1_fill if lv_cell.value == "Lv.1" else lv2_fill
            for col_idx in range(1, len(MASTER_COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.alignment = body_align
                cell.border = thin_border

        # --- Data validation on column X (Open_or_Close) ---
        # Column X is the 24th column
        dv = DataValidation(
            type="list",
            formula1='"O,X,O or X"',
            allow_blank=True,
        )
        dv.error = "O, X, 또는 'O or X'만 입력 가능합니다."
        dv.errorTitle = "입력 오류"
        dv.prompt = "O(공개), X(비공개), O or X(선택) 중 입력"
        dv.promptTitle = "Open or Close"
        # Apply to all data rows in column X
        dv.add(f"X2:X{len(df) + 1}")
        ws.add_data_validation(dv)

        # --- Freeze panes ---
        ws.freeze_panes = "D2"

    logger.info("Master Excel saved: %s (%d rows)", path, len(df))


# ---------------------------------------------------------------------------
# CLI / direct execution
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    base = "."
    if len(sys.argv) > 1:
        base = sys.argv[1]

    # Check --no-translate flag
    do_translate = "--no-translate" not in sys.argv

    result = build_master_sheet(base_dir=base, translate=do_translate)
    print(f"\nStatus: {result.status}")
    if result.data:
        for k, v in result.data.items():
            if k != "warnings":
                print(f"  {k}: {v}")
        if result.data.get("warnings"):
            print("  Warnings:")
            for w in result.data["warnings"]:
                print(f"    - {w}")
    if result.error_message:
        print(f"  Error: {result.error_message}")
    print(f"  Time: {result.processing_time_sec}s")
