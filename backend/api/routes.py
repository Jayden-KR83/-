"""FastAPI 라우터 - CDP AI Platform API"""
import shutil
import sqlite3
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
from fastapi import APIRouter, UploadFile, File, HTTPException, Query
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any

from backend.core.config import settings
from backend.core.skill_loader import list_available_skills
from backend.agents.pdf_parser_agent import run_pdf_parser, run_scoring_parser, PDF_TYPE_A, PDF_TYPE_B, PDF_TYPE_LABELS
from backend.agents.crawl_agent import run_crawl_agent

router = APIRouter(prefix="/api/v1")


# ─────────────────────────────────
# 시스템 상태
# ─────────────────────────────────
@router.get("/health")
def health_check():
    """서버 상태 및 설정 확인"""
    import socket
    config_errors = settings.validate()

    # 현재 서버 IP 목록 수집
    hostname = socket.gethostname()
    ips = []
    try:
        for info in socket.getaddrinfo(hostname, None):
            ip = info[4][0]
            if ip and not ip.startswith("127.") and ":" not in ip:
                if ip not in ips:
                    ips.append(ip)
    except Exception:
        pass

    return {
        "status": "ok" if not config_errors else "degraded",
        "claude_api": {
            "model": settings.CLAUDE_MODEL,
            "key_configured": bool(settings.ANTHROPIC_API_KEY),
            "key_preview": (settings.ANTHROPIC_API_KEY[:12] + "...") if settings.ANTHROPIC_API_KEY else "미설정",
        },
        "skills_loaded": list_available_skills(),
        "config_errors": config_errors,
        "server": {
            "hostname": hostname,
            "ips": ips,
            "access_urls": [f"http://{ip}:8000" for ip in ips] + [f"http://{hostname}:8000"],
        },
    }


# ─────────────────────────────────
# Selenium SPA Crawler (Chrome headless)
# ─────────────────────────────────
def _crawl_with_selenium(url: str, wait_sec: int = 5) -> str:
    """JavaScript SPA 사이트를 Chrome headless + Selenium으로 크롤링."""
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service

        opts = Options()
        opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--disable-dev-shm-usage")
        opts.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

        # ChromeDriver 경로 (프로젝트 내 또는 PATH)
        driver_path = Path("c:/Project/CDP-AI-Platform/data/chromedriver.exe")
        if driver_path.exists():
            service = Service(str(driver_path))
        else:
            service = Service()  # PATH에서 찾기

        driver = webdriver.Chrome(service=service, options=opts)
        driver.set_page_load_timeout(30)
        driver.get(url)

        import time as _t
        _t.sleep(wait_sec)

        text = driver.find_element("tag name", "body").text
        driver.quit()
        return text
    except Exception as e:
        import logging
        logging.getLogger("cdp.routes").warning("Selenium crawl failed: %s", e)
        return ""


# ─────────────────────────────────
# Parsing Progress Tracking
# ─────────────────────────────────
_parse_progress = {
    "running": False, "step": "", "file": "", "current": 0, "total": 0,
    "elapsed": 0, "done": False, "error": "",
}


@router.get("/parse-progress")
def get_parse_progress():
    """파싱 진행 상태 조회."""
    return _parse_progress


# ─────────────────────────────────
# PDF Parser Agent — 완전 분리된 엔드포인트
# ─────────────────────────────────
def _save_upload(file: UploadFile) -> Path:
    """업로드 파일 저장 + 크기 검증. 손상 시 knowledge에서 복구."""
    save_path = settings.UPLOAD_DIR / file.filename
    settings.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    file.file.seek(0)
    with open(save_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    if save_path.stat().st_size < 100:
        kb = Path("c:/Project/CDP-AI-Platform/data/knowledge") / file.filename
        if kb.exists() and kb.stat().st_size > 100:
            shutil.copy2(str(kb), str(save_path))
        else:
            raise HTTPException(400, f"파일 손상 ({save_path.stat().st_size} bytes). 다시 업로드해주세요.")
    return save_path


@router.post("/parse-pdf")
async def parse_pdf(
    file: UploadFile = File(...),
    page_start: Optional[int] = Query(None),
    page_end: Optional[int] = Query(None),
):
    """CDP PDF 업로드. 파일명에 'scoring' 포함 시 채점 파서, 아니면 문항 파서."""
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "PDF 파일만 업로드 가능합니다")
    save_path = _save_upload(file)

    # 파일명에 scoring 포함 → run_scoring_parser 직접 호출
    if "scoring" in file.filename.lower():
        result = run_scoring_parser(
            pdf_path=str(save_path),
            output_dir=str(settings.OUTPUT_DIR),
            save_excel=True,
            page_start=page_start,
            page_end=page_end,
        )
    else:
        result = run_pdf_parser(
            pdf_path=str(save_path),
            output_dir=str(settings.OUTPUT_DIR),
            save_excel=True,
            page_start=page_start,
            page_end=page_end,
            pdf_type=PDF_TYPE_A,
        )
    return result.model_dump()


@router.post("/parse-scoring")
async def parse_scoring(
    file: UploadFile = File(...),
    page_start: Optional[int] = Query(None),
    page_end: Optional[int] = Query(None),
):
    """Scoring Methodology PDF 전용 엔드포인트 (백업)."""
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "PDF 파일만 업로드 가능합니다")
    save_path = _save_upload(file)
    result = run_scoring_parser(
        pdf_path=str(save_path),
        output_dir=str(settings.OUTPUT_DIR),
        save_excel=True,
        page_start=page_start,
        page_end=page_end,
    )
    return result.model_dump()


# ─────────────────────────────────
# Workspace Reset + Step Status
# ─────────────────────────────────
@router.post("/reset-workspace")
def reset_workspace():
    """data/outputs/ 내 모든 파일 삭제 (knowledge는 유지). 초기 상태로 리셋."""
    import glob
    outputs_dir = Path("data/outputs")
    deleted = []
    if outputs_dir.exists():
        for f in outputs_dir.iterdir():
            if f.is_file() and f.name != ".gitkeep":
                f.unlink()
                deleted.append(f.name)
        # Also clean fresh/ subfolder
        fresh_dir = outputs_dir / "fresh"
        if fresh_dir.exists():
            for f in fresh_dir.iterdir():
                if f.is_file():
                    f.unlink()
                    deleted.append(f"fresh/{f.name}")
    return {"deleted": deleted, "count": len(deleted)}


@router.get("/step-status")
def get_step_status():
    """각 Step의 실행 상태를 반환 (결과물 존재 여부 기준)."""
    outputs = Path("data/outputs")
    master = outputs / "CDP_2025_Master_Analysis.xlsx"

    step1_a = list(outputs.glob("*Questionnaire*_parsed.xlsx")) if outputs.exists() else []
    step1_b = list(outputs.glob("*_scoring.xlsx")) if outputs.exists() else []
    has_master = master.exists() and master.stat().st_size > 1000

    # Check if master has answers (AF column)
    has_drafts = False
    if has_master:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(master), read_only=True)
            ws = wb.active
            ans_col = None
            for ci, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), 1):
                if cell.value == "Answer":
                    ans_col = ci
                    break
            if ans_col:
                for row in ws.iter_rows(min_row=2, max_row=10, min_col=ans_col, max_col=ans_col):
                    if row[0].value:
                        has_drafts = True
                        break
            wb.close()
        except Exception:
            pass

    return {
        "step1": {
            "status": "done" if (step1_a and step1_b) else ("partial" if (step1_a or step1_b) else "none"),
            "source_a": len(step1_a),
            "source_b": len(step1_b),
        },
        "step2": {
            "status": "done" if has_master else "none",
            "file": master.name if has_master else None,
        },
        "step3": {
            "status": "done" if has_drafts else "none",
        },
    }


# ─────────────────────────────────
# Crawl Agent
# ─────────────────────────────────
class CrawlRequest(BaseModel):
    urls: List[str] = []
    query: str = "CDP ESG 정보"
    use_playwright: bool = False
    summarize: bool = True
    category: Optional[str] = None  # "cdp_esg" | "rating_agencies" | "construction_esg"


@router.get("/crawl/presets")
def crawl_presets():
    """크롤링 카테고리 프리셋 목록 반환"""
    from backend.agents.crawl_agent import CRAWL_PRESETS
    return {
        k: {"label": v["label"], "url_count": len(v["urls"]), "urls": v["urls"]}
        for k, v in CRAWL_PRESETS.items()
    }


@router.post("/crawl")
def crawl(req: CrawlRequest):
    """ESG/CDP 관련 외부 URL 크롤링 + Claude 요약"""
    from backend.agents.crawl_agent import CRAWL_PRESETS
    if not req.urls and not req.category:
        raise HTTPException(status_code=400, detail="URL 목록 또는 category가 필요합니다")
    result = run_crawl_agent(
        urls=req.urls,
        query=req.query,
        use_playwright=req.use_playwright,
        summarize=req.summarize,
        category=req.category,
    )
    return result.model_dump()


# ─────────────────────────────────
# 파일 관리
# ─────────────────────────────────
@router.get("/outputs")
def list_outputs():
    """생성된 결과 파일 목록"""
    settings.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    files = [
        {"name": f.name, "size_kb": round(f.stat().st_size / 1024, 1)}
        for f in settings.OUTPUT_DIR.iterdir() if f.is_file()
    ]
    return {"files": files}


@router.get("/download/{filename}")
def download_file(filename: str):
    """결과 파일 다운로드"""
    file_path = settings.OUTPUT_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail=f"파일 없음: {filename}")
    return FileResponse(path=str(file_path), filename=filename)


# ============================================================
# Phase 2: Scoring Agent
# ============================================================
from backend.agents.scoring_agent import run_scoring_agent

class ScoreRequest(BaseModel):
    question_id: str
    question_text: str
    answer_text: str
    max_points: float
    reference_data: str = None

@router.post("/score")
def score_answer(req: ScoreRequest):
    """CDP 답변 자동 채점 (Phase 2)"""
    if not req.answer_text.strip():
        raise HTTPException(status_code=400, detail="답변 내용이 비어있습니다")
    if req.max_points <= 0:
        raise HTTPException(status_code=400, detail="배점은 0보다 커야 합니다")
    result = run_scoring_agent(
        question_id=req.question_id,
        question_text=req.question_text,
        answer_text=req.answer_text,
        max_points=req.max_points,
        reference_data=req.reference_data,
    )
    return result.model_dump()


# ============================================================
# Phase 3: Answer Draft Agent
# ============================================================
from backend.agents.answer_draft_agent import run_answer_draft_agent
from backend.core.vector_store import get_vector_store

class DraftRequest(BaseModel):
    question_id: str
    question_text: str
    max_points: float
    company_data: dict = {}

@router.post("/draft-answer")
def draft_answer(req: DraftRequest):
    """CDP 문항 답변 초안 자동 생성 (Phase 3)"""
    if not req.question_text.strip():
        raise HTTPException(status_code=400, detail="문항 내용이 비어있습니다")
    result = run_answer_draft_agent(
        question_id=req.question_id,
        question_text=req.question_text,
        max_points=req.max_points,
        company_data=req.company_data,
    )
    return result.model_dump()


# ============================================================
# Communication Tool: Master Analysis Sheet
# ============================================================
from backend.agents.communication_tool import build_master_sheet
from backend.agents.drafting_agent import run_drafting_agent, draft_progress

class MasterSheetRequest(BaseModel):
    translate: bool = False

@router.post("/build-master")
def api_build_master(req: MasterSheetRequest):
    """Source A + Source B를 병합하여 CDP_2025_Master_Analysis.xlsx 생성"""
    result = build_master_sheet(base_dir=".", translate=req.translate)
    return result.model_dump()

@router.post("/draft-answers")
def api_draft_answers():
    """2024 전년도 답변 기반으로 2025 Master Analysis AF열 초안 자동 작성."""
    import threading
    if draft_progress.get("running"):
        raise HTTPException(409, "이미 초안 작성 진행 중입니다.")
    thread = threading.Thread(target=lambda: run_drafting_agent(base_dir="."), daemon=True)
    thread.start()
    return {"status": "started", "message": "초안 작성이 백그라운드에서 시작되었습니다."}


@router.get("/draft-progress")
def api_draft_progress():
    """초안 작성 진행 상태 조회."""
    return draft_progress


@router.get("/master-preview")
def master_preview(limit: int = Query(default=2000, le=5000)):
    """Master Analysis 파일의 미리보기 데이터 반환 (JSON)"""
    import pandas as pd
    import numpy as np
    path = Path("data/outputs/CDP_2025_Master_Analysis.xlsx")
    if not path.exists():
        raise HTTPException(status_code=404, detail="Master 파일이 없습니다. Master 생성 버튼을 먼저 실행하세요.")
    try:
        df = pd.read_excel(str(path), engine="openpyxl", dtype=str)
        df = df.fillna("").replace([np.inf, -np.inf], "")
        # Ensure all values are JSON-serializable strings
        for col in df.columns:
            df[col] = df[col].astype(str).replace("nan", "")
        rows = df.head(limit).to_dict(orient="records")
        return {"rows": rows, "total": len(df)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Master 파일 읽기 실패: {str(e)}")

@router.get("/download-master")
def download_master():
    """생성된 Master Analysis 파일 다운로드"""
    path = Path("data/outputs/CDP_2025_Master_Analysis.xlsx")
    if not path.exists():
        raise HTTPException(status_code=404, detail="Master 파일이 아직 생성되지 않았습니다. /build-master를 먼저 실행하세요.")
    return FileResponse(
        str(path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="CDP_2025_Master_Analysis.xlsx",
    )


@router.post("/reference/upload")
async def upload_reference(file: UploadFile = File(...)):
    """Reference 문서를 RAG 벡터 스토어에 추가 (Phase 3)"""
    if not file.filename.lower().endswith((".txt", ".md")):
        raise HTTPException(status_code=400, detail=".txt 또는 .md 파일만 지원합니다")
    content = (await file.read()).decode("utf-8", errors="ignore")
    store = get_vector_store()
    ok = store.add_document(
        doc_id=file.filename,
        text=content,
        metadata={"source": file.filename, "type": "uploaded"},
    )
    if not ok:
        raise HTTPException(status_code=500, detail="문서 저장 실패")
    return {"message": f"{file.filename} 저장 완료", "total_docs": store.count()}


@router.get("/reference/status")
def reference_status():
    """RAG 벡터 스토어 상태 확인"""
    store = get_vector_store()
    return {
        "total_docs": store.count(),
        "chromadb_active": store.is_chromadb_active(),
    }




# ─────────────────────────────────
# PDF 다중 파일 동시 파싱
# ─────────────────────────────────
@router.post("/parse-pdf/multi")
async def parse_pdf_multi(
    files: List[UploadFile] = File(...),
):
    """여러 PDF 파일을 동시에 파싱하여 각각 Excel로 저장.
    - 최대 10개 파일, 총 100MB 이하
    - 파일별 독립 처리 (한 파일 실패해도 나머지 계속)
    - 결과: 파일별 상태, Excel 경로, 처리 시간 반환
    """
    if not files:
        raise HTTPException(status_code=400, detail="파일이 없습니다")
    if len(files) > 10:
        raise HTTPException(status_code=400, detail="한 번에 최대 10개 파일까지 가능합니다")

    # 1. 파일 저장
    saved_paths: List[str] = []
    for f in files:
        if not f.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail=f"{f.filename}: PDF 파일만 가능합니다")
        save_path = settings.UPLOAD_DIR / f.filename
        settings.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        f.file.seek(0)
        with open(save_path, "wb") as out:
            shutil.copyfileobj(f.file, out)
        saved_paths.append(str(save_path))

    total_size_mb = sum(Path(p).stat().st_size for p in saved_paths) / 1024 / 1024
    if total_size_mb > 100:
        raise HTTPException(status_code=400, detail=f"총 파일 크기 {total_size_mb:.1f}MB 초과 (최대 100MB)")

    # 2. 동시 파싱 (ThreadPoolExecutor)
    results: List[Dict[str, Any]] = []
    max_workers = min(len(saved_paths), 3)  # 최대 3개 동시 처리

    def parse_one(pdf_path: str) -> Dict[str, Any]:
        filename = Path(pdf_path).name
        fname_lower = filename.lower()
        t0 = time.time()
        try:
            # 파일명에 scoring 포함 → run_scoring_parser 직접 호출
            if "scoring" in fname_lower:
                result = run_scoring_parser(
                    pdf_path=pdf_path,
                    output_dir=str(settings.OUTPUT_DIR),
                    save_excel=True,
                )
            else:
                result = run_pdf_parser(
                    pdf_path=pdf_path,
                    output_dir=str(settings.OUTPUT_DIR),
                    save_excel=True,
                    pdf_type=PDF_TYPE_A,
                )
            elapsed = round(time.time() - t0, 1)
            data = result.data if isinstance(result.data, dict) else {}
            return {
                "file": filename,
                "status": result.status.value if hasattr(result.status, "value") else str(result.status),
                "excel_path": data.get("excel_path", ""),
                "excel_filename": data.get("excel_filename", ""),
                "stats": data.get("excel_stats", {}),
                "elapsed_sec": elapsed,
                "error": result.error_message or "",
            }
        except Exception as e:
            return {
                "file": filename,
                "status": "failed",
                "excel_path": "",
                "excel_filename": "",
                "stats": {},
                "elapsed_sec": round(time.time() - t0, 1),
                "error": str(e),
            }

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(parse_one, path): path for path in saved_paths}
        for future in as_completed(futures):
            results.append(future.result())

    # 3. 요약
    success = [r for r in results if r["status"] == "success"]
    failed  = [r for r in results if r["status"] != "success"]

    return {
        "total": len(results),
        "success": len(success),
        "failed": len(failed),
        "results": sorted(results, key=lambda x: x["file"]),
        "download_links": [
            f"/api/v1/download/{r['excel_filename']}"
            for r in success if r["excel_filename"]
        ],
    }

# ============================================================
# DB 조회 / 배치 파싱 엔드포인트
# ============================================================
from backend.core.database import (
    list_sessions, get_session, delete_session,
    list_questions, count_questions, list_scoring_results,
)

@router.get("/sessions")
def get_sessions(limit: int = 50):
    return {"sessions": list_sessions(limit=limit)}

@router.get("/sessions/{session_id}")
def get_session_detail(session_id: int):
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail=f"세션 없음: {session_id}")
    questions = list_questions(session_id=session_id, limit=1000)
    return {"session": session, "questions": questions}

@router.delete("/sessions/{session_id}")
def remove_session(session_id: int):
    if not delete_session(session_id):
        raise HTTPException(status_code=404, detail=f"세션 없음: {session_id}")
    return {"message": f"세션 {session_id} 삭제 완료"}

@router.get("/questions")
def get_questions(
    session_id: Optional[int] = None,
    question_id: Optional[str] = None,
    limit: int = 200,
    offset: int = 0,
):
    questions = list_questions(session_id=session_id, question_id=question_id, limit=limit, offset=offset)
    total = count_questions(session_id=session_id)
    return {"total": total, "count": len(questions), "questions": questions}

@router.get("/scoring-history")
def get_scoring_history(question_id: Optional[str] = None, limit: int = 100):
    return {"results": list_scoring_results(question_id=question_id, limit=limit)}

class BatchParseRequest(BaseModel):
    filename: str
    page_start: Optional[int] = None
    page_end:   Optional[int] = None

@router.post("/parse-pdf/batch")
def parse_pdf_batch(req: BatchParseRequest):
    save_path = settings.UPLOAD_DIR / req.filename
    if not save_path.exists():
        raise HTTPException(status_code=404, detail=f"파일 없음: {req.filename}. 먼저 /parse-pdf로 업로드하세요.")
    if req.page_start and req.page_end and req.page_start > req.page_end:
        raise HTTPException(status_code=400, detail="page_start가 page_end보다 클 수 없습니다")
    result = run_pdf_parser(
        pdf_path=str(save_path),
        output_dir=str(settings.OUTPUT_DIR),
        save_excel=True,
        page_start=req.page_start,
        page_end=req.page_end,
    )
    return result.model_dump()


# ============================================================
# Knowledge Base + Chat (RAG) Endpoints
# ============================================================
import anthropic
from backend.core.knowledge_processor import extract_text, SUPPORTED_EXTENSIONS
from backend.core.vector_store import get_knowledge_store

_DB_PATH = "c:/Project/CDP-AI-Platform/data/cdp_platform.db"
KNOWLEDGE_DIR = Path("c:/Project/CDP-AI-Platform/data/knowledge")


def _init_kb_db():
    """knowledge_chunks / knowledge_meta 테이블 초기화 (없으면 생성)"""
    conn = sqlite3.connect(_DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS knowledge_chunks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_name TEXT,
            chunk_text TEXT,
            chunk_index INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS knowledge_meta (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_name TEXT UNIQUE,
            display_name TEXT,
            summary TEXT,
            year TEXT,
            file_size INTEGER DEFAULT 0,
            chunk_count INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()


def _gen_summary(doc_name: str, chunks: list) -> str:
    """첫 3개 청크를 Claude로 요약 (실패 시 빈 문자열 반환)"""
    try:
        text_sample = "\n\n".join(c["text"] for c in chunks[:3])[:3000]
        client = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)
        msg = client.messages.create(
            model=settings.CLAUDE_MODEL,
            max_tokens=300,
            system="당신은 CDP 문서 요약 전문가입니다. 핵심 내용을 한국어로 3~4문장으로 요약하세요.",
            messages=[{"role": "user", "content": f"다음 문서({doc_name})를 요약하세요:\n\n{text_sample}"}],
        )
        return msg.content[0].text.strip() if msg.content else ""
    except Exception:
        return ""


def _search_chunks_sqlite(query: str, top_k: int = 5) -> list:
    """SQLite 키워드 기반 청크 검색 (ChromaDB 폴백)"""
    words = query.lower().split()
    conn = sqlite3.connect(_DB_PATH)
    rows = conn.execute("SELECT doc_name, chunk_text FROM knowledge_chunks").fetchall()
    conn.close()
    scored = []
    for doc_name, chunk_text in rows:
        score = sum(1 for w in words if w in chunk_text.lower())
        if score > 0:
            scored.append((score, doc_name, chunk_text))
    scored.sort(reverse=True)
    return scored[:top_k]


# OCR 진행 상태 추적
_ocr_status = {"running": False, "doc_name": "", "current_page": 0, "total_pages": 0,
               "ocr_pages": 0, "done": False, "error": "", "chunks": 0}


@router.get("/knowledge/ocr-status")
def get_ocr_status():
    """OCR 재처리 진행 상태 조회."""
    return _ocr_status


@router.post("/knowledge/reprocess-ocr")
def reprocess_knowledge_with_ocr(doc_name: str = Query(...)):
    """지식창고의 PDF 문서를 OCR 포함하여 재처리 (백그라운드)."""
    global _ocr_status
    _init_kb_db()
    file_path = KNOWLEDGE_DIR / doc_name
    if not file_path.exists():
        raise HTTPException(404, f"파일 없음: {doc_name}")
    if not doc_name.lower().endswith(".pdf"):
        raise HTTPException(400, "PDF 파일만 OCR 재처리 가능합니다")
    if _ocr_status["running"]:
        raise HTTPException(409, f"이미 OCR 진행 중: {_ocr_status['doc_name']} ({_ocr_status['current_page']}/{_ocr_status['total_pages']})")

    import threading

    def _run_ocr():
        global _ocr_status
        try:
            _ocr_status = {"running": True, "doc_name": doc_name, "current_page": 0,
                           "total_pages": 0, "ocr_pages": 0, "done": False, "error": "", "chunks": 0}

            # extract_text가 내부적으로 OCR을 호출하므로, 진행 상태를 전달하기 위해
            # _ocr_pdf_page를 래핑
            from backend.core import knowledge_processor as kp
            original_ocr = kp._ocr_pdf_page

            def _tracked_ocr(fp, pi):
                _ocr_status["ocr_pages"] += 1
                return original_ocr(fp, pi)

            kp._ocr_pdf_page = _tracked_ocr

            # 총 페이지 수 파악
            import pdfplumber
            with pdfplumber.open(str(file_path)) as pdf:
                _ocr_status["total_pages"] = len(pdf.pages)

            # 진행 상태를 위해 직접 추출
            chunks = []
            with pdfplumber.open(str(file_path)) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    _ocr_status["current_page"] = page_num
                    text = page.extract_text() or ""

                    try:
                        has_images = bool(page.images)
                    except Exception:
                        has_images = False

                    if len(text.strip()) < 200 or has_images:
                        ocr_text = _tracked_ocr(str(file_path), page_num - 1)
                        if ocr_text:
                            existing_words = set(text.lower().split())
                            new_parts = [w for w in ocr_text.split() if w.lower() not in existing_words]
                            if len(new_parts) > 10:
                                text = text + "\n[이미지 텍스트]\n" + ocr_text

                    from backend.core.knowledge_processor import _chunk_text
                    for chunk in _chunk_text(text):
                        chunks.append({"text": chunk, "page": page_num, "source": doc_name})

            kp._ocr_pdf_page = original_ocr  # restore

            if not chunks:
                _ocr_status["error"] = "텍스트 추출 실패"
                _ocr_status["running"] = False
                return

            store = get_knowledge_store()
            store.add_chunks(doc_name, chunks)

            conn = sqlite3.connect(_DB_PATH)
            conn.execute("DELETE FROM knowledge_chunks WHERE doc_name = ?", (doc_name,))
            for idx, chunk in enumerate(chunks):
                conn.execute(
                    "INSERT INTO knowledge_chunks (doc_name, chunk_text, chunk_index) VALUES (?, ?, ?)",
                    (doc_name, chunk["text"], idx),
                )
            conn.execute(
                "UPDATE knowledge_meta SET chunk_count = ? WHERE doc_name = ?",
                (len(chunks), doc_name),
            )
            conn.commit()
            conn.close()

            _ocr_status["chunks"] = len(chunks)
            _ocr_status["done"] = True
            _ocr_status["running"] = False

        except Exception as e:
            _ocr_status["error"] = str(e)
            _ocr_status["running"] = False

    thread = threading.Thread(target=_run_ocr, daemon=True)
    thread.start()

    return {"status": "started", "doc_name": doc_name, "message": "OCR 재처리가 백그라운드에서 시작되었습니다. 진행 상태는 상단에 표시됩니다."}


@router.post("/knowledge/upload")
async def upload_knowledge(files: List[UploadFile] = File(...)):
    """다양한 형식의 파일들을 지식베이스에 추가 — 텍스트 추출 후 ChromaDB 저장"""
    KNOWLEDGE_DIR.mkdir(parents=True, exist_ok=True)
    _init_kb_db()

    if not files:
        raise HTTPException(status_code=400, detail="파일이 없습니다")

    store = get_knowledge_store()
    docs = []

    for f in files:
        ext = Path(f.filename).suffix.lower()
        if ext not in SUPPORTED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail=f"{f.filename}: 지원하지 않는 파일 형식입니다. 지원 형식: {', '.join(SUPPORTED_EXTENSIONS.keys())}"
            )

        # 파일 저장
        save_path = KNOWLEDGE_DIR / f.filename
        with open(save_path, "wb") as out:
            shutil.copyfileobj(f.file, out)

        # 텍스트 추출
        try:
            chunks = extract_text(str(save_path))
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"{f.filename} 텍스트 추출 실패: {str(e)}")

        # ChromaDB에 저장
        added = store.add_chunks(f.filename, chunks)

        # SQLite에도 저장 (폴백 및 메타데이터용)
        file_size = save_path.stat().st_size
        chunk_count = added if added > 0 else len(chunks)
        summary = _gen_summary(f.filename, chunks)
        # 파일명에서 연도 추출 (예: cdp_guidance_2025.pdf → 2025)
        import re as _re
        year_m = _re.search(r'(20\d{2})', f.filename)
        year = year_m.group(1) if year_m else ""

        conn = sqlite3.connect(_DB_PATH)
        conn.execute("DELETE FROM knowledge_chunks WHERE doc_name = ?", (f.filename,))
        for idx, chunk in enumerate(chunks):
            conn.execute(
                "INSERT INTO knowledge_chunks (doc_name, chunk_text, chunk_index) VALUES (?, ?, ?)",
                (f.filename, chunk["text"], idx),
            )
        conn.execute("""
            INSERT INTO knowledge_meta (doc_name, display_name, summary, year, file_size, chunk_count)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(doc_name) DO UPDATE SET
                display_name = excluded.display_name,
                summary      = excluded.summary,
                year         = excluded.year,
                file_size    = excluded.file_size,
                chunk_count  = excluded.chunk_count
        """, (f.filename, f.filename, summary, year, file_size, chunk_count))
        conn.commit()
        conn.close()

        docs.append({
            "name": f.filename,
            "file_type": SUPPORTED_EXTENSIONS.get(ext, ext),
            "chunks_added": chunk_count,
            "summary": summary,
            "year": year,
            "status": "success",
        })

    return {"docs": docs, "total_docs": len(docs)}


class UrlKnowledgeRequest(BaseModel):
    url: str
    doc_name: Optional[str] = None  # 사용자 지정 이름 (미지정 시 URL에서 추출)


@router.post("/knowledge/url")
def add_knowledge_from_url(req: UrlKnowledgeRequest):
    """URL에서 웹페이지/PDF를 크롤링하여 지식창고에 원문 저장 (LLM 요약 없음, 토큰 비용 0).
    - HTML 페이지: 텍스트 추출 후 청크 분할
    - PDF 직접 링크 (.pdf): 다운로드 후 PDF 텍스트 추출
    """
    import re as _re
    from urllib.parse import urlparse

    url = req.url.strip()
    if not url.startswith("http"):
        raise HTTPException(400, "http:// 또는 https://로 시작하는 URL을 입력하세요.")

    KNOWLEDGE_DIR.mkdir(parents=True, exist_ok=True)
    _init_kb_db()

    parsed_url = urlparse(url)
    domain = parsed_url.netloc
    is_pdf_url = url.lower().endswith(".pdf")

    # 문서 이름 결정
    if req.doc_name:
        doc_name = req.doc_name
    elif is_pdf_url:
        doc_name = Path(parsed_url.path).name
    else:
        doc_name = f"url_{domain}_{Path(parsed_url.path).stem or 'index'}.txt"

    try:
        if is_pdf_url:
            # PDF 직접 다운로드
            import urllib.request as _ureq
            import ssl as _ssl
            _ctx = _ssl.create_default_context()
            _ctx.check_hostname = False
            _ctx.verify_mode = _ssl.CERT_NONE
            save_path = KNOWLEDGE_DIR / doc_name
            req = _ureq.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with _ureq.urlopen(req, timeout=30, context=_ctx) as resp:
                with open(save_path, "wb") as f:
                    f.write(resp.read())
            chunks = extract_text(str(save_path))
        else:
            # HTML 크롤링 (기본 크롤러 → SPA면 Selenium fallback)
            from backend.core.crawler import WebCrawler
            crawler = WebCrawler()
            result = crawler.fetch(url)
            content = result.content or ""
            title = result.title or doc_name

            # SPA 감지: 콘텐츠가 너무 짧거나 "JavaScript enabled" 포함
            if len(content) < 300 or "javascript enabled" in content.lower():
                # Selenium fallback 시도
                selenium_text = _crawl_with_selenium(url)
                if selenium_text and len(selenium_text) > 200:
                    content = selenium_text
                    title = doc_name
                else:
                    raise HTTPException(
                        400,
                        f"이 사이트는 JavaScript(SPA) 기반으로 자동 크롤링이 불가합니다. "
                        f"브라우저에서 해당 페이지를 열고 Ctrl+P → PDF로 저장 후, "
                        f"파일 업로드 기능을 이용해주세요. (URL: {url})"
                    )

            if not content or len(content.strip()) < 50:
                raise HTTPException(400, "페이지에서 충분한 텍스트를 추출하지 못했습니다.")

            # 텍스트를 청크로 분할 (500자 단위)
            CHUNK_SIZE = 500
            chunks = []
            for i in range(0, len(content), CHUNK_SIZE):
                chunk_text = content[i:i + CHUNK_SIZE].strip()
                if chunk_text:
                    chunks.append({"text": chunk_text, "page": i // CHUNK_SIZE + 1})

            # 텍스트 파일로도 저장
            save_path = KNOWLEDGE_DIR / doc_name
            with open(save_path, "w", encoding="utf-8") as f:
                f.write(f"# {title}\n# Source: {url}\n\n{content}")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"URL 처리 실패: {str(e)}")

    if not chunks:
        raise HTTPException(400, "텍스트를 추출하지 못했습니다.")

    # ChromaDB 저장
    store = get_knowledge_store()
    added = store.add_chunks(doc_name, chunks)

    # SQLite 저장 (원문, LLM 요약 없음)
    file_size = save_path.stat().st_size if save_path.exists() else 0
    chunk_count = added if added > 0 else len(chunks)
    summary = f"URL: {url} | {len(content) if not is_pdf_url else file_size} bytes | {chunk_count} chunks"

    year_m = _re.search(r"(20\d{2})", url + doc_name)
    year = year_m.group(1) if year_m else ""

    conn = sqlite3.connect(_DB_PATH)
    conn.execute("DELETE FROM knowledge_chunks WHERE doc_name = ?", (doc_name,))
    for idx, chunk in enumerate(chunks):
        conn.execute(
            "INSERT INTO knowledge_chunks (doc_name, chunk_text, chunk_index) VALUES (?, ?, ?)",
            (doc_name, chunk["text"], idx),
        )
    conn.execute("""
        INSERT INTO knowledge_meta (doc_name, display_name, summary, year, file_size, chunk_count)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(doc_name) DO UPDATE SET
            display_name = excluded.display_name,
            summary      = excluded.summary,
            year         = excluded.year,
            file_size    = excluded.file_size,
            chunk_count  = excluded.chunk_count
    """, (doc_name, doc_name, summary, year, file_size, chunk_count))
    conn.commit()
    conn.close()

    return {
        "doc_name": doc_name,
        "url": url,
        "type": "pdf" if is_pdf_url else "html",
        "chunks_added": chunk_count,
        "summary": summary,
        "status": "success",
    }


class ChatRequest(BaseModel):
    question: str
    mode: int = 2  # 1=KB전용, 2=KB+LLM, 3=LLM전용


@router.post("/chat")
def chat(req: ChatRequest):
    """Search-then-Answer 워크플로우:
    1) 지식베이스 검색
    2) 실시간 웹 검색 (필요 시)
    3) Claude 추론
    mode=1: 문서 한정, mode=2: 문서+AI (기본), mode=3: AI 추론
    """
    from backend.tools.web_search_tool import needs_web_search, search_cdp_sites, format_search_context, format_search_sources

    if not req.question.strip():
        raise HTTPException(status_code=400, detail="질문이 비어있습니다")

    mode = req.mode if req.mode in (1, 2, 3) else 2
    context = ""
    sources = []
    web_searched = False

    # ── Step 1: 지식베이스 검색 (mode 1, 2) ──────────────────
    # 한국어 질문 → 영어 키워드 추가 검색으로 영문 문서 매칭 강화
    _KO_EN_TERMS = {
        "탄소": "carbon", "공시": "disclosure", "기후": "climate", "변화": "change",
        "배출": "emission", "감축": "reduction", "목표": "target", "리스크": "risk",
        "전환": "transition", "에너지": "energy", "재생": "renewable", "점수": "score",
        "채점": "scoring", "방법론": "methodology", "질문": "question", "답변": "answer",
        "가이드": "guidance", "보고": "reporting", "경계": "boundary", "범위": "scope",
        "공급망": "supply chain", "가치사슬": "value chain", "중요성": "materiality",
        "거버넌스": "governance", "인센티브": "incentive", "물": "water", "산림": "forest",
        "생물다양성": "biodiversity", "폐기물": "waste", "정책": "policy",
    }

    search_queries = [req.question]
    # 한국어 키워드를 영어로 변환한 쿼리 추가
    en_words = []
    for ko, en in _KO_EN_TERMS.items():
        if ko in req.question:
            en_words.append(en)
    if en_words:
        search_queries.append(" ".join(en_words))

    if mode in (1, 2):
        _init_kb_db()
        store = get_knowledge_store()
        all_results = []
        for sq in search_queries:
            if store.is_active():
                results = store.search(sq, n_results=5)
                if results:
                    all_results.extend(results)
            if not all_results:
                sqlite_results = _search_chunks_sqlite(sq, top_k=5)
                if sqlite_results:
                    all_results.extend([{"text": c, "doc_name": d} for _, d, c in sqlite_results])

        # Deduplicate by text content
        seen = set()
        unique_results = []
        for r in all_results:
            txt_hash = hash(r["text"][:100])
            if txt_hash not in seen:
                seen.add(txt_hash)
                unique_results.append(r)

        if unique_results:
            context = "\n\n---\n\n".join([r["text"] for r in unique_results[:8]])
            sources = list({r["doc_name"] for r in unique_results[:8]})

        if mode == 1 and not context:
            return {"answer": "업로드된 지식베이스 문서에서 관련 내용을 찾을 수 없습니다.\n\n문서를 먼저 업로드하거나, 답변 모드를 변경해 주세요.", "sources": []}

    # ── Step 2: 실시간 웹 검색 (mode 2, 3 + 필요 시) ─────────
    if mode in (2, 3) and needs_web_search(req.question):
        try:
            web_results = search_cdp_sites(req.question)
            if web_results:
                web_context = format_search_context(web_results)
                web_sources = format_search_sources(web_results)
                context = (context + "\n\n" + web_context).strip() if context else web_context
                sources.extend(web_sources)
                web_searched = True
        except Exception as e:
            import logging
            logging.getLogger("cdp.chat").warning("Web search failed: %s", e)

    # ── Step 3: 시스템 프롬프트 (Multilingual + CDP Expert) ────
    base = (
        "당신은 SK에코플랜트의 CDP/ESG 전문 어시스턴트입니다.\n"
        "\n"
        "[역할]\n"
        "CDP(Carbon Disclosure Project) 질문지 응답, 채점 방법론(Disclosure/Awareness/Management/Leadership), "
        "기후변화 리스크 관리, GHG 배출량 산정, SBTi 목표, Scope 1/2/3 등에 깊은 전문 지식을 보유.\n"
        "\n"
        "[다국어 처리 규칙]\n"
        "1. 참고 자료가 영문이더라도 반드시 한국어로 답변 (단순 번역이 아닌 내용 해석)\n"
        "2. CDP 전문 용어는 '한국어(영어)' 형태로 병기: 예) 탄소 공시(Carbon Disclosure), "
        "이중 중요성(Double Materiality), 전환 계획(Transition Plan), "
        "과학 기반 감축 목표(Science Based Targets), 물리적 리스크(Physical Risk)\n"
        "3. 영문 가이드라인의 복잡한 내용은 한국 기업의 상황에 맞게 해석을 덧붙일 것\n"
        "4. 번역이 모호한 경우 원문을 일부 인용하며 설명: 예) 원문에서는 'financially material'이라 "
        "표현하는데, 이는 재무적으로 중대한 영향을 의미합니다.\n"
        "\n"
        "[답변 형식]\n"
        "1. 한국어로 답변\n"
        "2. 핵심만 간결하게 (불필요한 서론/맺음말 제거)\n"
        "3. 마크다운 기호(#, *, **, ---)는 절대 사용하지 말고 일반 텍스트와 줄바꿈만 사용\n"
        "4. 번호 목록은 '1.' 형식만 사용\n"
        "5. '사이트를 확인하세요'같은 회피성 답변 금지 — 구체적 날짜, 조건, 내용을 직접 답변\n"
        "6. 정보가 불확실하면 '(확인 필요)' 표시 후 가용 정보로 최대한 구체적 답변"
    )
    if mode == 1:
        system_prompt = base + "\n\n[모드: 문서 한정] 반드시 아래 참고 자료 내용만 근거로 답변하세요."
    elif mode == 2:
        system_prompt = base + "\n\n[모드: 문서+AI] 참고 자료를 우선 활용하고, 부족하면 CDP/ESG 전문 지식으로 보완하세요."
    else:
        system_prompt = base + "\n\n[모드: AI 추론] CDP/기후변화/ESG 전문 지식과 웹 검색 결과를 활용하여 답변하세요."

    # ── Step 4: LLM 호출 ─────────────────────────────────────
    if context:
        user_content = f"[참고 자료]\n{context[:4000]}\n\n[질문]\n{req.question}"
    else:
        user_content = req.question

    if not settings.ANTHROPIC_API_KEY:
        return {"answer": "API Key가 설정되지 않았습니다. .env 파일에 ANTHROPIC_API_KEY를 설정하세요.", "sources": [], "mode": mode}

    try:
        chat_model = "claude-haiku-4-5-20251001"
        client = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)
        msg = client.messages.create(
            model=chat_model,
            max_tokens=1500,
            system=system_prompt,
            messages=[{"role": "user", "content": user_content}],
        )
        answer = msg.content[0].text if msg.content else "답변을 생성할 수 없습니다."
    except anthropic.AuthenticationError:
        answer = "API Key 인증 실패. .env 파일의 ANTHROPIC_API_KEY를 확인하세요."
    except anthropic.RateLimitError:
        answer = "API 호출 한도 초과. 잠시 후 다시 시도하세요."
    except Exception as e:
        answer = f"AI 응답 오류: {str(e)}"

    return {"answer": answer, "sources": sources, "mode": mode, "web_searched": web_searched}


@router.get("/knowledge/docs")
def list_knowledge_docs():
    """지식 창고 전체 문서 목록 — 메타데이터(요약·연도·크기·청크 수) 포함"""
    _init_kb_db()
    conn = sqlite3.connect(_DB_PATH)
    rows = conn.execute("""
        SELECT doc_name, display_name, summary, year, file_size, chunk_count, created_at
        FROM knowledge_meta ORDER BY created_at DESC
    """).fetchall()
    conn.close()
    docs = []
    for doc_name, display_name, summary, year, file_size, chunk_count, created_at in rows:
        # 실제 파일 존재 여부 확인
        exists = (KNOWLEDGE_DIR / doc_name).exists()
        docs.append({
            "doc_name":     doc_name,
            "display_name": display_name or doc_name,
            "summary":      summary or "",
            "year":         year or "",
            "file_size_kb": round((file_size or 0) / 1024, 1),
            "chunk_count":  chunk_count or 0,
            "created_at":   created_at or "",
            "exists":       exists,
        })
    return {"docs": docs, "total": len(docs)}


class KnowledgeMetaUpdate(BaseModel):
    display_name: Optional[str] = None
    summary:      Optional[str] = None
    year:         Optional[str] = None


@router.put("/knowledge/meta/{doc_name:path}")
def update_knowledge_meta(doc_name: str, req: KnowledgeMetaUpdate):
    """지식 창고 문서 메타데이터(표시명·요약·연도) 수정"""
    _init_kb_db()
    conn = sqlite3.connect(_DB_PATH)
    row = conn.execute("SELECT id FROM knowledge_meta WHERE doc_name=?", (doc_name,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail=f"문서 없음: {doc_name}")
    fields, vals = [], []
    if req.display_name is not None:
        fields.append("display_name=?"); vals.append(req.display_name)
    if req.summary is not None:
        fields.append("summary=?");      vals.append(req.summary)
    if req.year is not None:
        fields.append("year=?");         vals.append(req.year)
    if fields:
        vals.append(doc_name)
        conn.execute(f"UPDATE knowledge_meta SET {', '.join(fields)} WHERE doc_name=?", vals)
        conn.commit()
    conn.close()
    return {"status": "updated", "doc_name": doc_name}


@router.get("/knowledge/list")
def list_knowledge():
    """업로드된 지식베이스 문서 목록 및 통계 반환"""
    store = get_knowledge_store()
    stats = store.get_stats()
    stats["supported_extensions"] = SUPPORTED_EXTENSIONS
    return stats


@router.get("/knowledge/stats")
def knowledge_stats():
    """지식베이스 통계 — ChromaDB 상태, 문서 수, 청크 수"""
    store = get_knowledge_store()
    stats = store.get_stats()
    stats["supported_formats"] = list(SUPPORTED_EXTENSIONS.keys())
    return stats


@router.delete("/knowledge/{doc_name:path}")
def delete_knowledge(doc_name: str):
    """지식베이스에서 문서 삭제 — ChromaDB + SQLite + 원본 파일"""
    store = get_knowledge_store()
    store.delete_document(doc_name)

    # SQLite에서도 삭제
    _init_kb_db()
    conn = sqlite3.connect(_DB_PATH)
    conn.execute("DELETE FROM knowledge_chunks WHERE doc_name = ?", (doc_name,))
    conn.execute("DELETE FROM knowledge_meta WHERE doc_name = ?", (doc_name,))
    conn.commit()
    conn.close()

    # 원본 파일 삭제
    file_path = KNOWLEDGE_DIR / doc_name
    if file_path.exists():
        file_path.unlink()

    return {"message": f"{doc_name} 삭제 완료"}


# ═══════════════════════════════════════════════════════════════════════════════
# CDP Auto Answer (답변 초안 자동 생성)
# ═══════════════════════════════════════════════════════════════════════════════

class CDPAutoAnswerRequest(BaseModel):
    excel_filename: str              # data/reference/ 에 업로드된 파일명
    sheet_name: str    = "Comm. Tool_SKEP"
    enable_ai: bool    = False       # False=Phase1만(무료), True=Phase1+2(API비용발생)
    output_mode: str   = "preview"   # "preview" | "write"
    target_questions: Optional[List[str]] = None   # 지정 시 해당 q_no만 처리
    company_context: str = ""


@router.get("/cdp-answer/excel-list")
def cdp_excel_list():
    """data/reference/ 폴더의 Excel 파일 목록 반환"""
    ref_dir = settings.REFERENCE_DIR
    ref_dir.mkdir(parents=True, exist_ok=True)
    files = [
        {"name": f.name, "size_kb": round(f.stat().st_size / 1024, 1)}
        for f in ref_dir.iterdir()
        if f.suffix.lower() in (".xlsx", ".xls")
    ]
    return {"files": files}


@router.post("/cdp-answer/verify-columns")
def cdp_verify_columns(req: CDPAutoAnswerRequest):
    """Excel 열 매핑 검증 — 실제 실행 전 반드시 먼저 호출"""
    from backend.agents.cdp_auto_answer_module import ExcelDataLoader
    excel_path = settings.REFERENCE_DIR / req.excel_filename
    if not excel_path.exists():
        raise HTTPException(status_code=404, detail=f"파일 없음: {req.excel_filename}")
    try:
        result = ExcelDataLoader.verify_columns(str(excel_path), req.sheet_name)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/cdp-answer/run")
def cdp_auto_answer_run(req: CDPAutoAnswerRequest):
    """
    CDP 1st Answer 자동 생성 실행.
    enable_ai=False: Phase 1만 (규칙 기반, 비용 없음)
    enable_ai=True : Phase 1 + 2 (AI 초안 포함, API 비용 발생)
    """
    from backend.agents.cdp_auto_answer_module import run_cdp_auto_answer
    excel_path = settings.REFERENCE_DIR / req.excel_filename
    if not excel_path.exists():
        raise HTTPException(status_code=404, detail=f"파일 없음: {req.excel_filename}")
    try:
        results = run_cdp_auto_answer(
            excel_path       = str(excel_path),
            sheet_name       = req.sheet_name,
            company_context  = req.company_context,
            enable_ai        = req.enable_ai,
            output_mode      = req.output_mode,
            target_questions = req.target_questions,
        )
        # questions 리스트는 크므로 summary만 포함하여 반환
        summary = {
            "total":      results["total"],
            "full_auto":  results["full_auto"],
            "ai_draft":   results["ai_draft"],
            "manual":     results["manual"],
            "skipped":    results["skipped"],
            "written":    results.get("written", 0),
            "auto_rate":  round((results["full_auto"] + results["skipped"]) / max(results["total"], 1) * 100, 1),
            "ai_rate":    round((results["full_auto"] + results["ai_draft"] + results["skipped"]) / max(results["total"], 1) * 100, 1),
        }
        # 검수 필요 항목 (상위 50개)
        review = [
            q for q in results["questions"]
            if q["needs_review"] and q["answer"]
        ][:50]
        # 완전 자동 처리 항목 샘플 (상위 20개)
        auto_sample = [
            q for q in results["questions"]
            if q["auto_level"] == "full_auto" and q["answer"]
        ][:20]
        return {
            "status":      "success",
            "summary":     summary,
            "review_list": review,
            "auto_sample": auto_sample,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/cdp-answer/upload-excel")
async def cdp_upload_excel(file: UploadFile = File(...)):
    """CDP Master Excel 파일을 data/reference/ 에 업로드"""
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Excel 파일만 업로드 가능합니다")
    ref_dir = settings.REFERENCE_DIR
    ref_dir.mkdir(parents=True, exist_ok=True)
    save_path = ref_dir / file.filename
    with open(save_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    size_kb = round(save_path.stat().st_size / 1024, 1)
    return {"filename": file.filename, "size_kb": size_kb, "path": str(save_path)}


# ── CDP 워크벤치 엔드포인트 (문항 조회 / 개별 생성 / 저장) ──────────────────

class CDPWorkbenchRequest(BaseModel):
    excel_filename: str
    sheet_name: str = "Comm. Tool_SKEP"

class CDPGenerateSingleRequest(BaseModel):
    excel_filename: str
    sheet_name: str = "Comm. Tool_SKEP"
    row_idx: int                        # Excel 행 번호 (1-based)
    company_context: str = ""

class CDPSaveAnswerRequest(BaseModel):
    excel_filename: str
    sheet_name: str = "Comm. Tool_SKEP"
    row_idx: int                        # Excel 행 번호 (1-based)
    answer: str                         # 저장할 답변
    column: str = "BX"                  # "BX"=1st_Ans, "CB"=Final_Ans


@router.post("/cdp-answer/questions")
def cdp_workbench_questions(req: CDPWorkbenchRequest):
    """
    CDP 워크벤치: 전체 문항 목록 반환.
    각 문항의 전체 데이터(질문/옵션/채점기준/기존답변/전년도답변) 포함.
    """
    from backend.agents.cdp_auto_answer_module import ExcelDataLoader, AnswerTypeClassifier
    excel_path = settings.REFERENCE_DIR / req.excel_filename
    if not excel_path.exists():
        raise HTTPException(status_code=404, detail=f"파일 없음: {req.excel_filename}")
    try:
        questions = ExcelDataLoader.load(str(excel_path), req.sheet_name)
        clf = AnswerTypeClassifier()
        result = []
        for q in questions:
            q.answer_type = clf.classify(q)
            d = q.to_dict()
            result.append(d)
        # 통계
        total    = len(result)
        open_q   = sum(1 for q in result if q["is_open"])
        answered = sum(1 for q in result if q["has_answer"])
        scored   = sum(1 for q in result if q["is_scored"])
        return {
            "status": "ok",
            "stats": {
                "total": total,
                "open":  open_q,
                "answered": answered,
                "unanswered": open_q - answered,
                "scored": scored,
            },
            "questions": result,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/cdp-answer/generate-single")
def cdp_generate_single(req: CDPGenerateSingleRequest):
    """
    CDP 워크벤치: 단일 문항 AI 초안 생성.
    row_idx로 해당 행을 특정, Claude API 호출 후 초안 반환 (저장하지 않음).
    할루시네이션 방지: BV 선택지 + W~AD 채점기준 + BZ 전년도답변만 참조.
    """
    from backend.agents.cdp_auto_answer_module import (
        ExcelDataLoader, AnswerTypeClassifier, AIAnswerGenerator,
        AnswerType, AutoLevel,
    )
    excel_path = settings.REFERENCE_DIR / req.excel_filename
    if not excel_path.exists():
        raise HTTPException(status_code=404, detail=f"파일 없음: {req.excel_filename}")
    try:
        questions = ExcelDataLoader.load(str(excel_path), req.sheet_name)
        q_map = {q.row_idx: q for q in questions}
        if req.row_idx not in q_map:
            raise HTTPException(status_code=404, detail=f"row_idx={req.row_idx} 없음")
        q = q_map[req.row_idx]
        clf = AnswerTypeClassifier()
        q.answer_type = clf.classify(q)
        if q.answer_type == AnswerType.SKIP:
            return {"status": "skipped", "reason": "BQ=X (Close 문항)", "answer": ""}
        if q.answer_type == AnswerType.ATTACHMENT:
            return {"status": "manual_required", "reason": "첨부 파일형 — 수동 작성 필요", "answer": ""}

        prompt  = AIAnswerGenerator.build_prompt(q, req.company_context)
        answer  = AIAnswerGenerator.call_claude_api(prompt)
        return {
            "status":       "ok",
            "row_idx":      q.row_idx,
            "q_no":         q.q_no,
            "sub_q_no":     q.sub_q_no,
            "answer_type":  q.answer_type.value,
            "answer":       answer,
            "prev_answer":  q.prev_year_answer,
            "scoring": {
                "dc": q.scoring_dc,
                "ac": q.scoring_ac,
                "mc": q.scoring_mc,
                "lc": q.scoring_lc,
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/cdp-answer/save-answer")
def cdp_save_answer(req: CDPSaveAnswerRequest):
    """
    CDP 워크벤치: 답변 Excel에 직접 저장.
    column="BX" → 2025 CDP 1st_Ans. (col index 76, 1-based)
    column="CB" → 2025 CDP_Final_Ans. (col index 81, 1-based)
    """
    from backend.agents.cdp_auto_answer_module import ExcelDataLoader
    import openpyxl

    excel_path = settings.REFERENCE_DIR / req.excel_filename
    if not excel_path.exists():
        raise HTTPException(status_code=404, detail=f"파일 없음: {req.excel_filename}")

    col_map = {"BX": 76, "CB": 81}   # 1-based column numbers
    if req.column not in col_map:
        raise HTTPException(status_code=400, detail=f"column은 BX 또는 CB만 허용")

    try:
        from backend.agents.cdp_auto_answer_module import ExcelDataLoader
        ExcelDataLoader._ensure_xlsx(str(excel_path))
        wb = openpyxl.load_workbook(str(excel_path))
        ws = wb[req.sheet_name]
        ws.cell(row=req.row_idx, column=col_map[req.column]).value = req.answer
        wb.save(str(excel_path))
        return {
            "status": "saved",
            "row_idx": req.row_idx,
            "column": req.column,
            "answer_preview": req.answer[:100] + ("..." if len(req.answer) > 100 else ""),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
