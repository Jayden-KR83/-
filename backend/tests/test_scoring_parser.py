# -*- coding: utf-8 -*-
"""Comprehensive test for Scoring Methodology parser."""
import json
import os
import tempfile

from backend.agents.pdf_parser_agent import (
    _build_scoring_questions,
    _detect_point_allocation_table,
    _extract_point_allocation,
    _save_scoring_excel,
    PDF_TYPE_A, PDF_TYPE_B,
    SCORING_HEADER_PATTERN, SCORING_QUESTION_REF_PATTERN,
    CRITERIA_HEADER_PATTERN, POINTS_MAX_PATTERN, ROUTE_PATTERN,
)


def test_patterns():
    """TEST 1: Pattern matching against real CDP text."""
    tests = [
        ("SCORING_HEADER", SCORING_HEADER_PATTERN, "1.5 - Scoring criteria", True),
        ("SCORING_HEADER", SCORING_HEADER_PATTERN, "2.1a - Scoring criteria", True),
        ("SCORING_HEADER", SCORING_HEADER_PATTERN, "Some random text", False),
        ("QUESTION_REF", SCORING_QUESTION_REF_PATTERN, "(1.5) - Provide details on your reporting boundary.", True),
        ("CRITERIA", CRITERIA_HEADER_PATTERN, "Disclosure criteria", True),
        ("CRITERIA", CRITERIA_HEADER_PATTERN, "Awareness criteria", True),
        ("CRITERIA", CRITERIA_HEADER_PATTERN, "Management criteria", True),
        ("CRITERIA", CRITERIA_HEADER_PATTERN, "Leadership criteria", True),
        ("MAX_PTS", POINTS_MAX_PATTERN, "A maximum of 2/2 points is available for this question.", True),
        ("MAX_PTS", POINTS_MAX_PATTERN, "A maximum of 0/0 points is available for this route.", True),
        ("ROUTE", ROUTE_PATTERN, "ROUTE A)", True),
        ("ROUTE", ROUTE_PATTERN, "ROUTE B)", True),
    ]
    for name, pat, text, should_match in tests:
        m = pat.search(text) if name == "MAX_PTS" else pat.match(text)
        matched = m is not None
        assert matched == should_match, f"{name} failed on '{text}'"
        print(f"  [PASS] {name}: '{text[:50]}'")


def _make_elements():
    """Build test elements simulating pages 12-13 from actual CDP scoring PDF screenshots."""
    return [
        {
            "type": "text", "page": 12, "y_top": 0,
            "data": (
                "1.5 - Scoring criteria\n"
                "Disclosure criteria\n"
                "Page 12 out of 797\n"
                "@cdp | www.cdp.net\n"
                "(1.5) - Provide details on your reporting boundary.\n"
                "Points will be awarded per completed cell in proportion to the number of cells displayed.\n"
                "A maximum of 2/2 points is available for this question.\n"
                "Awareness criteria\n"
                "Table completed - 1 point\n"
                "A maximum of 1/1 point is available for this question.\n"
                "Management criteria\n"
                "Full Awareness points must be awarded to be eligible for Management points.\n"
                "ROUTE A)\n"
                "Yes selected in column - 1 point\n"
                "A maximum of 1/1 point is available for this route.\n"
                "OR\n"
                "ROUTE B)\n"
                "Not applicable selected - 0/0 points\n"
                "A maximum of 0/0 points is available for this route.\n"
                "Leadership criteria\n"
                "Not scored."
            ),
        },
        {
            "type": "table", "page": 12, "y_top": 5,
            "data": [
                ["Theme", "Sector that scoring criteria apply to"],
                ["CC", "GN/CN/CO/CE/OG/EU/PF/FB/AC/CH/RE/MM/ST/TS/TO/TO-EPM/CG/FS"],
            ],
            "classified_rows": [
                {"row": ["Theme", "Sector that scoring criteria apply to"], "color_class": "je-mok"},
                {"row": ["CC", "GN/CN/CO/CE/OG/EU/PF/FB/AC/CH/RE/MM/ST/TS/TO/TO-EPM/CG/FS"], "color_class": "nae-yong"},
            ],
            "bbox": (0, 5, 500, 30),
        },
        {
            "type": "table", "page": 13, "y_top": 10,
            "data": [
                ["Disclosure numerator", "Disclosure denominator",
                 "Awareness numerator", "Awareness denominator",
                 "Management numerator", "Management denominator",
                 "Leadership numerator", "Leadership denominator"],
                ["2", "2", "1", "1", "0 or 1", "0 or 1", "0", "0"],
            ],
            "classified_rows": [],
            "bbox": (0, 10, 500, 50),
        },
    ]


def test_question_building():
    """TEST 2: Full question building from CDP text."""
    elements = _make_elements()
    questions = _build_scoring_questions(elements)

    assert len(questions) == 1, f"Expected 1 question, got {len(questions)}"
    q = questions[0]

    print(f"  ID: {q['문항ID']}")
    print(f"  Text: {q['질문내용'][:60]}")
    print(f"  Max pts: {q['최대배점']}")
    print(f"  D: {q['Disclosure_기준'][:40]}... | pts={q['Disclosure_배점']}")
    print(f"  A: {q['Awareness_기준'][:40]}... | pts={q['Awareness_배점']}")
    print(f"  M: {q['Management_기준'][:40]}... | pts={q['Management_배점']}")
    print(f"  L: {q['Leadership_기준']} | pts={q['Leadership_배점']}")
    print(f"  Alloc: D={q['D_num']}/{q['D_den']} A={q['A_num']}/{q['A_den']} M={q['M_num']}/{q['M_den']} L={q['L_num']}/{q['L_den']}")
    print(f"  Theme: {q['테마']} | Sector: {q['섹터'][:30]}...")

    assert q["문항ID"] == "1.5"
    assert "reporting boundary" in q["질문내용"].lower()
    assert q["최대배점"] == "2/2"
    assert q["Disclosure_배점"] == "2/2"
    assert q["Awareness_배점"]
    assert q["Leadership_기준"] == "Not scored."
    assert q["Leadership_배점"] == "0"
    assert q["D_num"] == "2"
    assert q["D_den"] == "2"
    assert q["L_num"] == "0"
    assert q["L_den"] == "0"
    assert q["섹터"]  # Should have sector data
    print("  [PASS] All question fields validated!")


def test_excel_save():
    """TEST 3: Excel save and readback."""
    elements = _make_elements()
    with tempfile.NamedTemporaryFile(suffix="_scoring.xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        stats = _save_scoring_excel(elements, tmp_path, source_file="test.pdf")
        print(f"  Stats: {json.dumps(stats)}")
        assert stats["structured_questions"] == 1

        fsize = os.path.getsize(tmp_path)
        print(f"  File size: {fsize} bytes")
        assert fsize > 1000, "Excel file too small"

        import openpyxl
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        print(f"  Headers: {headers[:6]}...")
        assert headers[0] == "문항ID", f"First header wrong: {headers[0]}"

        data_row = [c.value for c in ws[2]]
        print(f"  Row 2 first cell: {data_row[0]}")
        assert data_row[0] == "1.5", f"First data cell wrong: {data_row[0]}"
        print("  [PASS] Excel file verified!")
    finally:
        os.unlink(tmp_path)


def test_multi_question():
    """TEST 4: Multiple questions from different pages."""
    multi_elements = [
        {
            "type": "text", "page": 12, "y_top": 0,
            "data": (
                "1.5 - Scoring criteria\n"
                "(1.5) - Provide details on your reporting boundary.\n"
                "A maximum of 2/2 points is available for this question.\n"
                "Disclosure criteria\n"
                "Points awarded per completed cell - 2 points\n"
                "Awareness criteria\n"
                "Table completed - 1 point\n"
                "A maximum of 1/1 point is available for this question.\n"
                "Management criteria\n"
                "ROUTE A)\n"
                "Yes selected - 1 point\n"
                "A maximum of 1/1 point is available for this route.\n"
                "Leadership criteria\n"
                "Not scored."
            ),
        },
        {
            "type": "text", "page": 14, "y_top": 0,
            "data": (
                "2.1 - Scoring criteria\n"
                "(2.1) - Indicate whether your organization undertook a risk assessment.\n"
                "A maximum of 4/4 points is available for this question.\n"
                "Disclosure criteria\n"
                "Table completed - 2 points\n"
                "Awareness criteria\n"
                "Relevant risks identified - 1 point\n"
                "A maximum of 1/1 point is available for this question.\n"
                "Management criteria\n"
                "Risk management process described - 1 point\n"
                "A maximum of 1/1 point is available for this route.\n"
                "Leadership criteria\n"
                "Comprehensive approach demonstrated - 1 point\n"
                "A maximum of 1/1 point is available for this question."
            ),
        },
    ]
    questions = _build_scoring_questions(multi_elements)
    assert len(questions) == 2, f"Expected 2 questions, got {len(questions)}"
    assert questions[0]["문항ID"] == "1.5"
    assert questions[1]["문항ID"] == "2.1"
    assert questions[0]["Leadership_기준"] == "Not scored."
    assert "Comprehensive" in questions[1]["Leadership_기준"]
    print(f"  Q1: {questions[0]['문항ID']} - {questions[0]['질문내용'][:40]}...")
    print(f"  Q2: {questions[1]['문항ID']} - {questions[1]['질문내용'][:40]}...")
    print("  [PASS] Multi-question parsing correct!")


def test_point_allocation_table():
    """TEST 5: Point allocation table detection and extraction."""
    table_data = [
        ["Disclosure numerator", "Disclosure denominator",
         "Awareness numerator", "Awareness denominator",
         "Management numerator", "Management denominator",
         "Leadership numerator", "Leadership denominator"],
        ["2", "2", "1", "1", "0 or 1", "0 or 1", "0", "0"],
    ]
    assert _detect_point_allocation_table(table_data), "Detection failed"
    alloc = _extract_point_allocation(table_data)
    assert alloc["D_num"] == "2"
    assert alloc["D_den"] == "2"
    assert alloc["L_num"] == "0"
    assert alloc["L_den"] == "0"
    assert alloc["M_num"] == "0 or 1"
    print(f"  Allocation: {alloc}")
    print("  [PASS] Point allocation extraction correct!")


if __name__ == "__main__":
    print("=" * 60)
    print("TEST 1: Pattern matching")
    print("=" * 60)
    test_patterns()

    print()
    print("=" * 60)
    print("TEST 2: Question building")
    print("=" * 60)
    test_question_building()

    print()
    print("=" * 60)
    print("TEST 3: Excel save")
    print("=" * 60)
    test_excel_save()

    print()
    print("=" * 60)
    print("TEST 4: Multi-question")
    print("=" * 60)
    test_multi_question()

    print()
    print("=" * 60)
    print("TEST 5: Point Allocation table")
    print("=" * 60)
    test_point_allocation_table()

    print()
    print("=" * 60)
    print("ALL 5 TESTS PASSED")
    print("=" * 60)
